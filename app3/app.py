import os
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, make_response
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import re
import base64
from werkzeug.utils import secure_filename

# Initialize Flask app
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-here-change-in-production')

# ============ DATABASE CONFIGURATION ============
# Use PostgreSQL on Render, SQLite locally
if os.environ.get('DATABASE_URL'):
    # Render PostgreSQL
    database_url = os.environ.get('DATABASE_URL')
    # Fix for Render's PostgreSQL URL format
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'pool_pre_ping': True,
        'pool_recycle': 300,
        'pool_size': 5,
        'max_overflow': 10
    }
    print("✅ Using PostgreSQL database on Render")
else:
    # Local SQLite
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///prepayment.db'
    print("✅ Using local SQLite database")

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create upload folder if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Initialize extensions
from app3.models import db, Company, PrepaymentSchedule, AmortizationEntry
db.init_app(app)


# ============ SAFE DATABASE INITIALIZATION ============
def init_database():
    """Initialize database tables and create default company - SAFE VERSION"""
    try:
        # Create all tables if they don't exist
        db.create_all()
        print("✅ Database tables verified/created successfully!")
    except Exception as e:
        print(f"⚠️ Table verification: {e}")
        print("✅ Database may already be initialized.")

    # Create default company ONLY if none exists
    try:
        if not Company.query.first():
            company = Company(
                name='Default Company',
                code='DEF001',
                address='123 Business St',
                phone='555-0000',
                email='default@company.com',
                is_active=True
            )
            db.session.add(company)
            db.session.commit()
            print("✅ Default company created!")
        else:
            # Just log that data exists, don't modify anything
            company_count = Company.query.count()
            schedule_count = PrepaymentSchedule.query.count()
            print(f"✅ Database already has data: {company_count} companies, {schedule_count} schedules")
    except Exception as e:
        print(f"⚠️ Company check: {e}")
        print("✅ Database may already have data. Skipping initialization.")


# ============ RUN DATABASE INITIALIZATION ON STARTUP ============
with app.app_context():
    init_database()
    print("=" * 50)
    print("🚀 Database initialized successfully!")
    print(f"📊 Database: {app.config['SQLALCHEMY_DATABASE_URI']}")
    print("=" * 50)


# ============ VALIDATION FUNCTIONS ============
def validate_account_number(value, field_name):
    """
    Validate account numbers - allows:
    - Letters (A-Z, a-z)
    - Numbers (0-9)
    - Spaces
    - Hyphens (-)
    - Underscores (_)
    - Periods (.)
    - Slashes (/)
    - Plus signs (+)
    - Parentheses ()
    - Square brackets []
    - Curly braces {}
    - Colon (:)
    - Semicolon (;)
    - Ampersand (&)
    - At symbol (@)
    - Exclamation (!)
    - Question mark (?)
    - Apostrophe (')
    - Comma (,)
    """
    if not value:
        return False, f"{field_name} is required"

    # Allow alphanumeric, spaces, and common special characters
    pattern = r'^[a-zA-Z0-9\s\-_\.\/\+\(\)\[\]\{\}:;&@!?\'",]+$'

    if not re.match(pattern, value):
        return False, f"{field_name} contains invalid characters. Allowed: letters, numbers, spaces, -, _, ., /, +, (), [], {{}}, :, ;, &, @, !, ?, ', ,"

    if len(value) < 2:
        return False, f"{field_name} must be at least 2 characters long"

    if len(value) > 200:
        return False, f"{field_name} must be less than 200 characters"

    return True, ""


def validate_positive_number(value, field_name):
    """Validate that a value is a positive number"""
    try:
        num = float(value)
        if num <= 0:
            return False, f"{field_name} must be greater than zero"
        return True, ""
    except ValueError:
        return False, f"{field_name} must be a valid number"


def validate_positive_integer(value, field_name):
    """Validate that a value is a positive integer"""
    try:
        num = int(value)
        if num <= 0:
            return False, f"{field_name} must be greater than zero"
        return True, ""
    except ValueError:
        return False, f"{field_name} must be a valid whole number"


def validate_date(value, field_name):
    """Validate date format"""
    try:
        datetime.strptime(value, '%Y-%m-%d')
        return True, ""
    except ValueError:
        return False, f"{field_name} must be in YYYY-MM-DD format"


def format_number(num):
    """Format number with comma separators and 2 decimal places"""
    if num is None:
        return "0.00"
    return f"{num:,.2f}"


# ============ CONTEXT PROCESSOR ============
@app.context_processor
def inject_company():
    company_name = None
    company_id = session.get('current_company')
    if company_id:
        company = Company.query.get(company_id)
        if company:
            company_name = company.name
    return dict(current_company_name=company_name, format_number=format_number)


# ============ SETUP DEFAULT COMPANY ============
def setup_default_company():
    """Create default company if none exists"""
    with app.app_context():
        if not Company.query.first():
            company = Company(
                name='Default Company',
                code='DEF001',
                address='123 Business St',
                phone='555-0000',
                email='default@company.com',
                is_active=True
            )
            db.session.add(company)
            db.session.commit()
            print("✅ Default company created!")
            return company
    return None


# ============ REPORT HELPER FUNCTIONS ============
def get_monthly_amortization(schedule, year, month):
    """Get amortization amount for a specific month"""
    entries = AmortizationEntry.query.filter_by(schedule_id=schedule.id).all()
    total = 0
    for entry in entries:
        if entry.due_date.year == year and entry.due_date.month == month:
            total += float(entry.amount)
    return total


def get_opening_balance(schedule, start_date):
    """Get opening balance (remaining prepayment) at the start date"""
    if schedule.amortize_start_period >= start_date:
        return 0.0

    entries = AmortizationEntry.query.filter(
        AmortizationEntry.schedule_id == schedule.id,
        AmortizationEntry.due_date < start_date
    ).all()

    amortized_before = sum(float(e.amount) for e in entries)
    opening = float(schedule.total_cost) - amortized_before

    if opening < 0:
        opening = 0

    return opening


def get_period_additions(schedule, start_date, end_date):
    """Get additions (new prepayments) in the period based on amortize start period"""
    if start_date <= schedule.amortize_start_period <= end_date:
        return float(schedule.total_cost)
    return 0


def get_closing_balance(schedule, end_date):
    """Get closing balance after end date"""
    entries = AmortizationEntry.query.filter(
        AmortizationEntry.schedule_id == schedule.id,
        AmortizationEntry.due_date <= end_date
    ).all()
    total_amortized = sum(float(e.amount) for e in entries)
    return float(schedule.total_cost) - total_amortized


def get_schedules_by_debit_account(schedules):
    """Group schedules by debit account"""
    grouped = {}
    for schedule in schedules:
        debit = schedule.debit_account
        if debit not in grouped:
            grouped[debit] = []
        grouped[debit].append(schedule)
    return grouped


def filter_schedules_for_report(company_id, start_date, end_date):
    """Common function to filter schedules for reports"""
    all_schedules = PrepaymentSchedule.query.filter_by(company_id=company_id).all()
    schedules = []

    for schedule in all_schedules:
        entries = AmortizationEntry.query.filter_by(schedule_id=schedule.id).order_by(
            AmortizationEntry.due_date).all()

        include_schedule = False

        if entries:
            if schedule.amortize_start_period <= end_date:
                for entry in entries:
                    if start_date <= entry.due_date <= end_date:
                        include_schedule = True
                        break

                if not include_schedule:
                    if schedule.amortize_start_period < start_date:
                        for entry in entries:
                            if entry.due_date >= start_date:
                                include_schedule = True
                                break
                    elif schedule.amortize_start_period == start_date:
                        include_schedule = True
                    elif schedule.amortize_start_period > start_date and schedule.amortize_start_period <= end_date:
                        include_schedule = True

        if include_schedule:
            schedules.append(schedule)

    return schedules


# ============ UTILITY FUNCTIONS ============
def generate_excel_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    headers = ['Debit Account', 'Credit Account', 'Transaction Date', 'Description',
               'Total Cost', 'Period to Amortize', 'Amortize Start Period']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = 25

    examples = [
        ['6020/000 Prepayment Middleware Integration', '2010/000 Accounts Payable', '2026-01-01',
         'Software License Prepayment', '5000.00', '12', '2026-01-01'],
        ['6030/001 Prepayment Insurance', '2010/001 Insurance Payable', '2026-01-15', 'Annual Insurance Premium',
         '2400.00', '12', '2026-01-15'],
        ['1010/000 Cash Account', '6020/000 Prepayment Middleware Integration', '2026-02-01', 'Payment for Services',
         '1200.00', '6', '2026-02-01']
    ]

    for row_num, row_data in enumerate(examples, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    note_row = len(examples) + 4
    ws.cell(row=note_row, column=1,
            value="Note: Account codes can contain numbers, letters, spaces, hyphens, underscores, periods, and slashes (/)")
    ws.cell(row=note_row, column=1).font = Font(italic=True, color='FF0000')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=7)

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


def import_schedules_from_excel(excel_file, company_id):
    """Import schedules from Excel with better error handling"""
    wb = None
    ws = None
    imported_count = 0
    error_count = 0

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"❌ Error loading Excel file: {str(e)}")
        return 0

    print("=" * 50)
    print("DEBUG: Excel Import Started")
    print("=" * 50)

    headers = []
    for col in range(1, 8):
        cell_value = ws.cell(row=1, column=col).value
        headers.append(cell_value)
    print(f"Headers found: {headers}")

    total_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            total_rows += 1
    print(f"Total data rows found: {total_rows}")
    print("=" * 50)

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        if not row[0]:
            print(f"Row {row_num}: Skipping - empty first column")
            continue

        try:
            if not all(row[:7]):
                missing = []
                if not row[0]: missing.append("Debit Account")
                if not row[1]: missing.append("Credit Account")
                if not row[2]: missing.append("Transaction Date")
                if not row[3]: missing.append("Description")
                if not row[4]: missing.append("Total Cost")
                if not row[5]: missing.append("Period to Amortize")
                if not row[6]: missing.append("Amortize Start Period")
                print(f"Row {row_num}: Skipping - missing columns: {', '.join(missing)}")
                error_count += 1
                continue

            debit_account = str(row[0]).strip()
            credit_account = str(row[1]).strip()
            transaction_date_value = row[2]
            description = str(row[3]).strip()
            total_cost_value = row[4]
            period_to_amortize_value = row[5]
            amortize_start_period_value = row[6]

            print(f"\nRow {row_num}: Processing...")
            print(f"  Debit: {debit_account}")
            print(f"  Credit: {credit_account}")

            valid, msg = validate_account_number(debit_account, 'Debit Account')
            if not valid:
                print(f"  ❌ Debit Account validation failed: {msg}")
                error_count += 1
                continue

            valid, msg = validate_account_number(credit_account, 'Credit Account')
            if not valid:
                print(f"  ❌ Credit Account validation failed: {msg}")
                error_count += 1
                continue

            try:
                if isinstance(transaction_date_value, datetime):
                    transaction_date = transaction_date_value.date()
                elif isinstance(transaction_date_value, date):
                    transaction_date = transaction_date_value
                else:
                    date_str = str(transaction_date_value).strip()
                    if ' ' in date_str:
                        date_str = date_str.split(' ')[0]
                    transaction_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                print(f"  ✅ Parsed transaction date: {transaction_date}")
            except Exception as e:
                print(f"  ❌ Invalid transaction date: {transaction_date_value} - Error: {str(e)}")
                error_count += 1
                continue

            try:
                if isinstance(amortize_start_period_value, datetime):
                    amortize_start_period = amortize_start_period_value.date()
                elif isinstance(amortize_start_period_value, date):
                    amortize_start_period = amortize_start_period_value
                else:
                    date_str = str(amortize_start_period_value).strip()
                    if ' ' in date_str:
                        date_str = date_str.split(' ')[0]
                    amortize_start_period = datetime.strptime(date_str, '%Y-%m-%d').date()
                print(f"  ✅ Parsed start period: {amortize_start_period}")
            except Exception as e:
                print(f"  ❌ Invalid amortize start period: {amortize_start_period_value} - Error: {str(e)}")
                error_count += 1
                continue

            try:
                if isinstance(total_cost_value, (int, float)):
                    total_cost = float(total_cost_value)
                else:
                    cost_str = str(total_cost_value).strip()
                    cost_str = cost_str.replace(',', '').replace('$', '').replace('£', '').replace('€', '')
                    total_cost = float(cost_str)

                if total_cost <= 0:
                    print(f"  ❌ Total cost must be positive: {total_cost}")
                    error_count += 1
                    continue
                print(f"  ✅ Parsed total cost: {total_cost}")
            except Exception as e:
                print(f"  ❌ Invalid total cost: {total_cost_value} - Error: {str(e)}")
                error_count += 1
                continue

            try:
                if isinstance(period_to_amortize_value, (int, float)):
                    period_to_amortize = int(period_to_amortize_value)
                else:
                    period_str = str(period_to_amortize_value).strip()
                    period_to_amortize = int(float(period_str))

                if period_to_amortize <= 0:
                    print(f"  ❌ Period to amortize must be positive: {period_to_amortize}")
                    error_count += 1
                    continue
                print(f"  ✅ Parsed period to amortize: {period_to_amortize}")
            except Exception as e:
                print(f"  ❌ Invalid period to amortize: {period_to_amortize_value} - Error: {str(e)}")
                error_count += 1
                continue

            schedule = PrepaymentSchedule(
                company_id=company_id,
                created_by_id=1,
                debit_account=debit_account,
                credit_account=credit_account,
                transaction_date=transaction_date,
                description=description,
                total_cost=total_cost,
                period_to_amortize=period_to_amortize,
                amortize_start_period=amortize_start_period
            )
            db.session.add(schedule)
            db.session.flush()

            for entry_data in schedule.get_amortization_schedule():
                entry = AmortizationEntry(
                    schedule_id=schedule.id,
                    period=entry_data['period'],
                    due_date=entry_data['date'],
                    amount=entry_data['amount'],
                    remaining_balance=entry_data['remaining_balance'],
                    status='PENDING'
                )
                db.session.add(entry)

            imported_count += 1
            print(f"  ✅ Successfully imported row {row_num}")

        except Exception as e:
            print(f"  ❌ Error on row {row_num}: {str(e)}")
            error_count += 1
            continue

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"❌ Database commit error: {str(e)}")
        return 0

    try:
        if wb:
            wb.close()
    except:
        pass

    print("\n" + "=" * 50)
    print("IMPORT SUMMARY:")
    print(f"  ✅ Successfully imported: {imported_count}")
    print(f"  ❌ Errors: {error_count}")
    print(f"  📊 Total rows processed: {imported_count + error_count}")
    print("=" * 50)

    return imported_count


def export_schedules_to_excel(schedules):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedules"

    headers = ['ID', 'Debit Account', 'Credit Account', 'Transaction Date', 'Description',
               'Total Cost', 'Periods', 'Remaining Balance']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws.column_dimensions[get_column_letter(col)].width = 22

    for row_num, schedule in enumerate(schedules, 2):
        ws.cell(row=row_num, column=1, value=schedule.id)
        ws.cell(row=row_num, column=2, value=schedule.debit_account)
        ws.cell(row=row_num, column=3, value=schedule.credit_account)
        ws.cell(row=row_num, column=4, value=schedule.transaction_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=5, value=schedule.description[:50])
        ws.cell(row=row_num, column=6, value=float(schedule.total_cost))
        ws.cell(row=row_num, column=7, value=schedule.period_to_amortize)
        ws.cell(row=row_num, column=8, value=float(schedule.remaining_balance))

    total_row = len(schedules) + 3
    ws.cell(row=total_row, column=5, value="TOTAL:")
    ws.cell(row=total_row, column=6, value=sum(float(s.total_cost) for s in schedules))
    ws.cell(row=total_row, column=8, value=sum(float(s.remaining_balance) for s in schedules))

    for col in range(1, 9):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


def generate_monthly_report_excel(schedules, start_date, end_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Report"

    months = []
    current = start_date
    while current <= end_date:
        months.append(current.strftime('%b'))
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)

    grouped_schedules = get_schedules_by_debit_account(schedules)

    title_cols = len(months) + 6
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_cols)
    title_cell = ws.cell(row=1, column=1,
                         value=f"Monthly Prepayment Report by Debit Account: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    headers = ['#', 'Transaction Date', 'Description', 'Opening', 'Period Additions'] + months + ['Closing']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = 18 if col <= 3 else 15

    row_num = 3
    grand_totals = {'opening': 0, 'additions': 0, 'closing': 0}
    grand_monthly_totals = {month: 0 for month in months}

    for debit_account, account_schedules in grouped_schedules.items():
        account_header_cell = ws.cell(row=row_num, column=1, value=debit_account)
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=title_cols)
        account_header_cell.font = Font(bold=True, size=12, color='FFFFFF')
        account_header_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        account_header_cell.alignment = Alignment(horizontal='left', vertical='center')
        row_num += 1

        account_totals = {'opening': 0, 'additions': 0, 'closing': 0}
        account_monthly_totals = {month: 0 for month in months}

        for idx, schedule in enumerate(account_schedules, 1):
            opening = get_opening_balance(schedule, start_date)
            account_totals['opening'] += opening
            grand_totals['opening'] += opening

            additions = get_period_additions(schedule, start_date, end_date)
            account_totals['additions'] += additions
            grand_totals['additions'] += additions

            monthly_amounts = []
            current_date = start_date
            for month in months:
                amount = get_monthly_amortization(schedule, current_date.year, current_date.month)
                monthly_amounts.append(amount)
                account_monthly_totals[month] += amount
                grand_monthly_totals[month] += amount
                if current_date.month == 12:
                    current_date = current_date.replace(year=current_date.year + 1, month=1)
                else:
                    current_date = current_date.replace(month=current_date.month + 1)

            closing = get_closing_balance(schedule, end_date)
            account_totals['closing'] += closing
            grand_totals['closing'] += closing

            ws.cell(row=row_num, column=1, value=idx)
            ws.cell(row=row_num, column=2, value=schedule.transaction_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=3, value=schedule.description)
            ws.cell(row=row_num, column=4, value=round(opening, 2))
            ws.cell(row=row_num, column=5, value=round(additions, 2))

            col = 6
            for amount in monthly_amounts:
                ws.cell(row=row_num, column=col, value=round(amount, 2))
                col += 1

            ws.cell(row=row_num, column=col, value=round(closing, 2))
            row_num += 1

        ws.cell(row=row_num, column=1, value="Subtotal")
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
        ws.cell(row=row_num, column=4, value=round(account_totals['opening'], 2))
        ws.cell(row=row_num, column=5, value=round(account_totals['additions'], 2))

        col = 6
        for month in months:
            ws.cell(row=row_num, column=col, value=round(account_monthly_totals[month], 2))
            col += 1

        ws.cell(row=row_num, column=col, value=round(account_totals['closing'], 2))

        for col in range(1, title_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
            cell.alignment = Alignment(horizontal='center' if col <= 3 else 'right', vertical='center')

        row_num += 1

    ws.cell(row=row_num, column=1, value="GRAND TOTAL")
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
    ws.cell(row=row_num, column=4, value=round(grand_totals['opening'], 2))
    ws.cell(row=row_num, column=5, value=round(grand_totals['additions'], 2))

    col = 6
    for month in months:
        ws.cell(row=row_num, column=col, value=round(grand_monthly_totals[month], 2))
        col += 1

    ws.cell(row=row_num, column=col, value=round(grand_totals['closing'], 2))

    for col in range(1, title_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        cell.alignment = Alignment(horizontal='center' if col <= 3 else 'right', vertical='center')

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=2, max_row=row_num, min_col=1, max_col=title_cols):
        for cell in row:
            cell.border = border
            if cell.row > 2:
                if cell.column > 3:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                else:
                    cell.alignment = Alignment(horizontal='left' if cell.column == 3 else 'center', vertical='center')

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


def generate_monthly_report_pdf(schedules, start_date, end_date):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16,
                                 textColor=colors.HexColor('#366092'), spaceAfter=15)
    story.append(Paragraph('Monthly Prepayment Report by Debit Account', title_style))

    date_style = ParagraphStyle('DateStyle', parent=styles['Normal'], fontSize=10,
                                textColor=colors.HexColor('#666666'), spaceAfter=12)
    story.append(
        Paragraph(f'Period: {start_date.strftime("%B %d, %Y")} to {end_date.strftime("%B %d, %Y")}', date_style))
    story.append(Spacer(1, 12))

    months = []
    current = start_date
    while current <= end_date:
        months.append(current.strftime('%b'))
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)

    grouped_schedules = get_schedules_by_debit_account(schedules)

    grand_totals = {'opening': 0, 'additions': 0, 'closing': 0}
    grand_monthly_totals = {month: 0 for month in months}

    table_data = []

    headers = ['#', 'Transaction Date', 'Description', 'Opening', 'Period Additions'] + months + ['Closing']
    table_data.append(headers)

    for debit_account, account_schedules in grouped_schedules.items():
        account_header = [debit_account] + [''] * (len(headers) - 1)
        table_data.append(account_header)

        account_totals = {'opening': 0, 'additions': 0, 'closing': 0}
        account_monthly_totals = {month: 0 for month in months}

        for idx, schedule in enumerate(account_schedules, 1):
            opening = get_opening_balance(schedule, start_date)
            account_totals['opening'] += opening
            grand_totals['opening'] += opening

            additions = get_period_additions(schedule, start_date, end_date)
            account_totals['additions'] += additions
            grand_totals['additions'] += additions

            monthly_amounts = []
            current_date = start_date
            for month in months:
                amount = get_monthly_amortization(schedule, current_date.year, current_date.month)
                monthly_amounts.append(amount)
                account_monthly_totals[month] += amount
                grand_monthly_totals[month] += amount
                if current_date.month == 12:
                    current_date = current_date.replace(year=current_date.year + 1, month=1)
                else:
                    current_date = current_date.replace(month=current_date.month + 1)

            closing = get_closing_balance(schedule, end_date)
            account_totals['closing'] += closing
            grand_totals['closing'] += closing

            row = [
                      str(idx),
                      schedule.transaction_date.strftime('%Y-%m-%d'),
                      schedule.description[:25] + '...' if len(schedule.description) > 25 else schedule.description,
                      format_number(opening),
                      format_number(additions)
                  ] + [format_number(amt) for amt in monthly_amounts] + [format_number(closing)]
            table_data.append(row)

        subtotal_row = ['Subtotal', '', '',
                        format_number(account_totals['opening']),
                        format_number(account_totals['additions'])] + \
                       [format_number(account_monthly_totals[month]) for month in months] + \
                       [format_number(account_totals['closing'])]
        table_data.append(subtotal_row)

    grand_total_row = ['GRAND TOTAL', '', '',
                       format_number(grand_totals['opening']),
                       format_number(grand_totals['additions'])] + \
                      [format_number(grand_monthly_totals[month]) for month in months] + \
                      [format_number(grand_totals['closing'])]
    table_data.append(grand_total_row)

    col_widths = [0.4 * inch, 0.9 * inch, 2.0 * inch, 0.9 * inch, 1.0 * inch] + [0.7 * inch] * len(months) + [
        0.9 * inch]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]

    for i, row in enumerate(table_data):
        if row[0] and not row[1] and not row[2] and i > 0:
            style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#4472C4')))
            style.append(('TEXTCOLOR', (0, i), (-1, i), colors.white))
            style.append(('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold'))
            style.append(('FONTSIZE', (0, i), (-1, i), 8))
            style.append(('ALIGN', (0, i), (-1, i), 'LEFT'))
        elif row[0] and row[0] == 'Subtotal':
            style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#B4C6E7')))
            style.append(('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold'))
            style.append(('FONTSIZE', (0, i), (-1, i), 7))
        elif row[0] and row[0] == 'GRAND TOTAL':
            style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FFC000')))
            style.append(('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold'))
            style.append(('FONTSIZE', (0, i), (-1, i), 8))

    for i in range(len(table_data)):
        for j in range(3, len(headers)):
            style.append(('ALIGN', (j, i), (j, i), 'RIGHT'))

    table.setStyle(TableStyle(style))
    story.append(table)

    story.append(Spacer(1, 12))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#666666'))
    story.append(Paragraph(f'Generated on: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}', footer_style))
    story.append(Paragraph(f'Total Records: {len(schedules)}', footer_style))

    doc.build(story)
    buffer.seek(0)
    return buffer


# ============ ROUTES ============

@app.route('/')
def index():
    if not session.get('current_company'):
        company = Company.query.first()
        if company:
            session['current_company'] = company.id
        else:
            company = setup_default_company()
            if company:
                session['current_company'] = company.id
    return redirect(url_for('dashboard'))


@app.route('/dashboard')
def dashboard():
    company_id = session.get('current_company')
    if not company_id:
        company = Company.query.first()
        if company:
            session['current_company'] = company.id
            company_id = company.id
        else:
            setup_default_company()
            company = Company.query.first()
            if company:
                session['current_company'] = company.id
                company_id = company.id

    schedules = PrepaymentSchedule.query.filter_by(company_id=company_id).all()
    total_cost = sum(float(s.total_cost) for s in schedules)
    remaining = sum(float(s.remaining_balance) for s in schedules)

    upcoming = AmortizationEntry.query.join(PrepaymentSchedule).filter(
        PrepaymentSchedule.company_id == company_id,
        AmortizationEntry.due_date >= date.today(),
        AmortizationEntry.status == 'PENDING'
    ).order_by(AmortizationEntry.due_date).limit(10).all()

    recent = PrepaymentSchedule.query.filter_by(company_id=company_id).order_by(
        PrepaymentSchedule.created_at.desc()
    ).limit(10).all()

    company = Company.query.get(company_id)

    return render_template('dashboard.html',
                           total_schedules=len(schedules),
                           total_cost=total_cost,
                           remaining_balance=remaining,
                           upcoming_entries=upcoming,
                           recent_schedules=recent,
                           company_name=company.name if company else 'No Company')


@app.route('/schedules')
def schedule_list():
    company_id = session.get('current_company')
    if not company_id:
        company = Company.query.first()
        if company:
            session['current_company'] = company.id
            company_id = company.id

    query = PrepaymentSchedule.query.filter_by(company_id=company_id)
    search = request.args.get('search', '')
    if search:
        query = query.filter(
            (PrepaymentSchedule.description.contains(search)) |
            (PrepaymentSchedule.debit_account.contains(search)) |
            (PrepaymentSchedule.credit_account.contains(search))
        )

    schedules = query.order_by(PrepaymentSchedule.created_at.desc()).all()

    page = request.args.get('page', 1, type=int)
    per_page = 10
    total = len(schedules)
    start = (page - 1) * per_page
    end = start + per_page
    page_obj = schedules[start:end]

    return render_template('schedule_list.html',
                           page_obj=page_obj,
                           search_query=search,
                           page=page,
                           total_pages=(total + per_page - 1) // per_page)


@app.route('/schedule/create', methods=['GET', 'POST'])
def schedule_create():
    company_id = session.get('current_company')
    if not company_id:
        company = Company.query.first()
        if company:
            session['current_company'] = company.id
            company_id = company.id

    if request.method == 'POST':
        try:
            debit_account = request.form.get('debit_account', '').strip()
            credit_account = request.form.get('credit_account', '').strip()
            transaction_date = request.form.get('transaction_date', '').strip()
            description = request.form.get('description', '').strip()
            total_cost = request.form.get('total_cost', '').strip()
            period_to_amortize = request.form.get('period_to_amortize', '').strip()
            amortize_start_period = request.form.get('amortize_start_period', '').strip()

            errors = []

            valid, msg = validate_account_number(debit_account, 'Debit Account')
            if not valid:
                errors.append(msg)

            valid, msg = validate_account_number(credit_account, 'Credit Account')
            if not valid:
                errors.append(msg)

            if not transaction_date:
                errors.append('Transaction Date is required')
            else:
                valid, msg = validate_date(transaction_date, 'Transaction Date')
                if not valid:
                    errors.append(msg)

            if not description:
                errors.append('Description is required')
            elif len(description) < 3:
                errors.append('Description must be at least 3 characters long')

            valid, msg = validate_positive_number(total_cost, 'Total Cost')
            if not valid:
                errors.append(msg)

            valid, msg = validate_positive_integer(period_to_amortize, 'Period to Amortize')
            if not valid:
                errors.append(msg)

            if not amortize_start_period:
                errors.append('Amortize Start Period is required')
            else:
                valid, msg = validate_date(amortize_start_period, 'Amortize Start Period')
                if not valid:
                    errors.append(msg)

            if errors:
                for error in errors:
                    flash(f'❌ {error}', 'danger')
                return render_template('schedule_form.html', action='Create')

            schedule = PrepaymentSchedule(
                company_id=company_id,
                created_by_id=1,
                debit_account=debit_account,
                credit_account=credit_account,
                transaction_date=datetime.strptime(transaction_date, '%Y-%m-%d').date(),
                description=description,
                total_cost=float(total_cost),
                period_to_amortize=int(period_to_amortize),
                amortize_start_period=datetime.strptime(amortize_start_period, '%Y-%m-%d').date()
            )
            db.session.add(schedule)
            db.session.flush()

            for entry_data in schedule.get_amortization_schedule():
                entry = AmortizationEntry(
                    schedule_id=schedule.id,
                    period=entry_data['period'],
                    due_date=entry_data['date'],
                    amount=entry_data['amount'],
                    remaining_balance=entry_data['remaining_balance'],
                    status='PENDING'
                )
                db.session.add(entry)

            db.session.commit()
            flash('✅ Schedule created successfully!', 'success')
            return redirect(url_for('schedule_detail', pk=schedule.id))

        except ValueError as e:
            flash(f'❌ Invalid input: {str(e)}', 'danger')
        except Exception as e:
            flash(f'❌ Error: {str(e)}', 'danger')
            db.session.rollback()

    return render_template('schedule_form.html', action='Create')


@app.route('/schedule/<int:pk>')
def schedule_detail(pk):
    company_id = session.get('current_company')
    if not company_id:
        company = Company.query.first()
        if company:
            session['current_company'] = company.id
            company_id = company.id

    schedule = PrepaymentSchedule.query.filter_by(id=pk, company_id=company_id).first_or_404()
    amortization = schedule.get_amortization_schedule()
    return render_template('schedule_detail.html', schedule=schedule, amortization_schedule=amortization)


@app.route('/schedule/<int:pk>/update', methods=['GET', 'POST'])
def schedule_update(pk):
    company_id = session.get('current_company')
    if not company_id:
        company = Company.query.first()
        if company:
            session['current_company'] = company.id
            company_id = company.id

    schedule = PrepaymentSchedule.query.filter_by(id=pk, company_id=company_id).first_or_404()

    if request.method == 'POST':
        try:
            debit_account = request.form.get('debit_account', '').strip()
            credit_account = request.form.get('credit_account', '').strip()
            transaction_date = request.form.get('transaction_date', '').strip()
            description = request.form.get('description', '').strip()
            total_cost = request.form.get('total_cost', '').strip()
            period_to_amortize = request.form.get('period_to_amortize', '').strip()
            amortize_start_period = request.form.get('amortize_start_period', '').strip()

            errors = []

            valid, msg = validate_account_number(debit_account, 'Debit Account')
            if not valid:
                errors.append(msg)

            valid, msg = validate_account_number(credit_account, 'Credit Account')
            if not valid:
                errors.append(msg)

            if not transaction_date:
                errors.append('Transaction Date is required')
            else:
                valid, msg = validate_date(transaction_date, 'Transaction Date')
                if not valid:
                    errors.append(msg)

            if not description:
                errors.append('Description is required')
            elif len(description) < 3:
                errors.append('Description must be at least 3 characters long')

            valid, msg = validate_positive_number(total_cost, 'Total Cost')
            if not valid:
                errors.append(msg)

            valid, msg = validate_positive_integer(period_to_amortize, 'Period to Amortize')
            if not valid:
                errors.append(msg)

            if not amortize_start_period:
                errors.append('Amortize Start Period is required')
            else:
                valid, msg = validate_date(amortize_start_period, 'Amortize Start Period')
                if not valid:
                    errors.append(msg)

            if errors:
                for error in errors:
                    flash(f'❌ {error}', 'danger')
                return render_template('schedule_form.html', schedule=schedule, action='Update')

            schedule.debit_account = debit_account
            schedule.credit_account = credit_account
            schedule.transaction_date = datetime.strptime(transaction_date, '%Y-%m-%d').date()
            schedule.description = description
            schedule.total_cost = float(total_cost)
            schedule.period_to_amortize = int(period_to_amortize)
            schedule.amortize_start_period = datetime.strptime(amortize_start_period, '%Y-%m-%d').date()

            AmortizationEntry.query.filter_by(schedule_id=schedule.id).delete()

            for entry_data in schedule.get_amortization_schedule():
                entry = AmortizationEntry(
                    schedule_id=schedule.id,
                    period=entry_data['period'],
                    due_date=entry_data['date'],
                    amount=entry_data['amount'],
                    remaining_balance=entry_data['remaining_balance'],
                    status='PENDING'
                )
                db.session.add(entry)

            db.session.commit()
            flash('✅ Schedule updated successfully!', 'success')
            return redirect(url_for('schedule_detail', pk=schedule.id))

        except ValueError as e:
            flash(f'❌ Invalid input: {str(e)}', 'danger')
        except Exception as e:
            flash(f'❌ Error: {str(e)}', 'danger')
            db.session.rollback()

    return render_template('schedule_form.html', schedule=schedule, action='Update')


@app.route('/schedule/<int:pk>/delete', methods=['POST'])
def schedule_delete(pk):
    company_id = session.get('current_company')
    if not company_id:
        company = Company.query.first()
        if company:
            session['current_company'] = company.id
            company_id = company.id

    schedule = PrepaymentSchedule.query.filter_by(id=pk, company_id=company_id).first_or_404()
    try:
        db.session.delete(schedule)
        db.session.commit()
        flash('✅ Schedule deleted successfully!', 'success')
    except Exception as e:
        flash(f'❌ Error deleting: {str(e)}', 'danger')
        db.session.rollback()
    return redirect(url_for('schedule_list'))


@app.route('/schedule/delete-multiple', methods=['POST'])
def schedule_delete_multiple():
    """Delete multiple schedules at once"""
    company_id = session.get('current_company')
    if not company_id:
        flash('Please select a company first.', 'warning')
        return redirect(url_for('schedule_list'))

    schedule_ids = request.form.getlist('schedule_ids')

    if not schedule_ids:
        flash('No schedules selected for deletion.', 'warning')
        return redirect(url_for('schedule_list'))

    try:
        # Convert to integers
        ids = [int(id) for id in schedule_ids if id]

        # Delete schedules belonging to the current company
        deleted_count = PrepaymentSchedule.query.filter(
            PrepaymentSchedule.id.in_(ids),
            PrepaymentSchedule.company_id == company_id
        ).delete(synchronize_session=False)

        db.session.commit()
        flash(f'✅ Successfully deleted {deleted_count} schedules!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error deleting schedules: {str(e)}', 'danger')

    return redirect(url_for('schedule_list'))


@app.route('/export/excel')
def export_excel():
    company_id = session.get('current_company')
    if not company_id:
        company = Company.query.first()
        if company:
            session['current_company'] = company.id
            company_id = company.id

    schedules = PrepaymentSchedule.query.filter_by(company_id=company_id).all()
    if not schedules:
        flash('No schedules to export.', 'warning')
        return redirect(url_for('schedule_list'))

    try:
        excel_file = export_schedules_to_excel(schedules)
        response = make_response(
            send_file(excel_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
        response.headers['Content-Disposition'] = f'attachment; filename=schedules_{date.today()}.xlsx'
        return response
    except Exception as e:
        flash(f'❌ Error exporting: {str(e)}', 'danger')
        return redirect(url_for('schedule_list'))


@app.route('/import/excel', methods=['GET', 'POST'])
def import_excel():
    company_id = session.get('current_company')
    if not company_id:
        company = Company.query.first()
        if company:
            session['current_company'] = company.id
            company_id = company.id

    if request.method == 'POST':
        if 'excel_file' not in request.files:
            flash('No file uploaded.', 'danger')
            return redirect(url_for('import_excel'))

        file = request.files['excel_file']
        if file.filename == '':
            flash('No file selected.', 'danger')
            return redirect(url_for('import_excel'))

        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls)', 'danger')
            return redirect(url_for('import_excel'))

        temp_path = None
        try:
            temp_path = f"temp_import_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            file.save(temp_path)
            count = import_schedules_from_excel(temp_path, company_id)

            if count > 0:
                flash(f'✅ Successfully imported {count} schedules!', 'success')
            else:
                flash('⚠️ No schedules were imported. Please check the file format.', 'warning')

        except Exception as e:
            flash(f'❌ Error importing: {str(e)}', 'danger')

        finally:
            if temp_path and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except PermissionError:
                    import time
                    time.sleep(1)
                    try:
                        os.remove(temp_path)
                    except:
                        pass

        return redirect(url_for('schedule_list'))

    return render_template('import_excel.html')


@app.route('/download/template')
def download_template():
    try:
        excel_file = generate_excel_template()
        response = make_response(
            send_file(excel_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
        response.headers['Content-Disposition'] = 'attachment; filename=prepayment_template.xlsx'
        return response
    except Exception as e:
        flash(f'❌ Error downloading template: {str(e)}', 'danger')
        return redirect(url_for('import_excel'))


# ============ DEBIT ACCOUNT REPORT ROUTES ============

@app.route('/report', methods=['GET', 'POST'])
def generate_report():
    company_id = session.get('current_company')
    if not company_id:
        company = Company.query.first()
        if company:
            session['current_company'] = company.id
            company_id = company.id

    return render_template('report_form.html')


@app.route('/report/view', methods=['GET', 'POST'])
def view_report():
    company_id = session.get('current_company')
    if not company_id:
        company = Company.query.first()
        if company:
            session['current_company'] = company.id
            company_id = company.id

    if request.method == 'POST':
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
    else:
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

    if not start_date_str or not end_date_str:
        flash('Please select both start and end dates.', 'danger')
        return redirect(url_for('generate_report'))

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        if start_date > end_date:
            flash('Start date must be before end date.', 'danger')
            return redirect(url_for('generate_report'))

        schedules = filter_schedules_for_report(company_id, start_date, end_date)

        if not schedules:
            flash('No schedules found for this period.', 'warning')
            return redirect(url_for('generate_report'))

        months = []
        current = start_date
        while current <= end_date:
            months.append(current.strftime('%b'))
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)

        session['report_start_date'] = start_date_str
        session['report_end_date'] = end_date_str

        grouped_schedules = get_schedules_by_debit_account(schedules)

        report_groups = []
        grand_totals = {'opening': 0, 'additions': 0, 'closing': 0}
        grand_monthly_totals = {month: 0 for month in months}

        for debit_account, account_schedules in grouped_schedules.items():
            account_data = []
            account_totals = {'opening': 0, 'additions': 0, 'closing': 0}
            account_monthly_totals = {month: 0 for month in months}

            for schedule in account_schedules:
                opening = get_opening_balance(schedule, start_date)
                account_totals['opening'] += opening
                grand_totals['opening'] += opening

                additions = get_period_additions(schedule, start_date, end_date)
                account_totals['additions'] += additions
                grand_totals['additions'] += additions

                monthly_amounts = []
                current_date = start_date
                for month in months:
                    amount = get_monthly_amortization(schedule, current_date.year, current_date.month)
                    monthly_amounts.append(amount)
                    account_monthly_totals[month] += amount
                    grand_monthly_totals[month] += amount
                    if current_date.month == 12:
                        current_date = current_date.replace(year=current_date.year + 1, month=1)
                    else:
                        current_date = current_date.replace(month=current_date.month + 1)

                closing = get_closing_balance(schedule, end_date)
                account_totals['closing'] += closing
                grand_totals['closing'] += closing

                account_data.append({
                    'transaction_date': schedule.transaction_date.strftime('%Y-%m-%d'),
                    'description': schedule.description,
                    'opening': opening,
                    'additions': additions,
                    'monthly_amounts': monthly_amounts,
                    'closing': closing
                })

            report_groups.append({
                'debit_account': debit_account,
                'schedules': account_data,
                'totals': account_totals,
                'monthly_totals': account_monthly_totals
            })

        return render_template('report_results.html',
                               report_groups=report_groups,
                               months=months,
                               grand_totals=grand_totals,
                               grand_monthly_totals=grand_monthly_totals,
                               start_date=start_date,
                               end_date=end_date,
                               generated_on=datetime.now(),
                               format_number=format_number)

    except Exception as e:
        flash(f'❌ Error generating report: {str(e)}', 'danger')
        return redirect(url_for('generate_report'))


@app.route('/report/download/pdf')
def download_report_pdf():
    start_date_str = request.args.get('start_date') or session.get('report_start_date')
    end_date_str = request.args.get('end_date') or session.get('report_end_date')

    if not start_date_str or not end_date_str:
        flash('Please generate the report first.', 'warning')
        return redirect(url_for('generate_report'))

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        company_id = session.get('current_company')

        schedules = filter_schedules_for_report(company_id, start_date, end_date)

        if not schedules:
            flash('No schedules found for this period.', 'warning')
            return redirect(url_for('generate_report'))

        pdf_file = generate_monthly_report_pdf(schedules, start_date, end_date)
        response = make_response(send_file(pdf_file, mimetype='application/pdf'))
        response.headers['Content-Disposition'] = f'attachment; filename=monthly_report_{start_date}_to_{end_date}.pdf'
        return response

    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'danger')
        return redirect(url_for('generate_report'))


@app.route('/report/download/excel')
def download_report_excel():
    start_date_str = request.args.get('start_date') or session.get('report_start_date')
    end_date_str = request.args.get('end_date') or session.get('report_end_date')

    if not start_date_str or not end_date_str:
        flash('Please generate the report first.', 'warning')
        return redirect(url_for('generate_report'))

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        company_id = session.get('current_company')

        schedules = filter_schedules_for_report(company_id, start_date, end_date)

        if not schedules:
            flash('No schedules found for this period.', 'warning')
            return redirect(url_for('generate_report'))

        excel_file = generate_monthly_report_excel(schedules, start_date, end_date)
        response = make_response(
            send_file(excel_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
        response.headers['Content-Disposition'] = f'attachment; filename=monthly_report_{start_date}_to_{end_date}.xlsx'
        return response

    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'danger')
        return redirect(url_for('generate_report'))


# ============ CREDIT ACCOUNT REPORT HELPER FUNCTIONS ============
def get_schedules_by_credit_account(schedules):
    """Group schedules by credit account"""
    grouped = {}
    for schedule in schedules:
        credit = schedule.credit_account
        if credit not in grouped:
            grouped[credit] = []
        grouped[credit].append(schedule)
    return grouped


def generate_credit_report_excel(schedules, start_date, end_date):
    """Generate Excel report with horizontal months grouped by credit account"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Credit Report"

    months = []
    current = start_date
    while current <= end_date:
        months.append(current.strftime('%b'))
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)

    grouped_schedules = get_schedules_by_credit_account(schedules)

    title_cols = len(months) + 6
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_cols)
    title_cell = ws.cell(row=1, column=1,
                         value=f"Monthly Prepayment Report by Credit Account: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    headers = ['#', 'Transaction Date', 'Description', 'Opening', 'Period Additions'] + months + ['Closing']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')  # Green theme
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = 18 if col <= 3 else 15

    row_num = 3
    grand_totals = {'opening': 0, 'additions': 0, 'closing': 0}
    grand_monthly_totals = {month: 0 for month in months}

    for credit_account, account_schedules in grouped_schedules.items():
        account_header_cell = ws.cell(row=row_num, column=1, value=credit_account)
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=title_cols)
        account_header_cell.font = Font(bold=True, size=12, color='FFFFFF')
        account_header_cell.fill = PatternFill(start_color='388E3C', end_color='388E3C', fill_type='solid')
        account_header_cell.alignment = Alignment(horizontal='left', vertical='center')
        row_num += 1

        account_totals = {'opening': 0, 'additions': 0, 'closing': 0}
        account_monthly_totals = {month: 0 for month in months}

        for idx, schedule in enumerate(account_schedules, 1):
            opening = get_opening_balance(schedule, start_date)
            account_totals['opening'] += opening
            grand_totals['opening'] += opening

            additions = get_period_additions(schedule, start_date, end_date)
            account_totals['additions'] += additions
            grand_totals['additions'] += additions

            monthly_amounts = []
            current_date = start_date
            for month in months:
                amount = get_monthly_amortization(schedule, current_date.year, current_date.month)
                monthly_amounts.append(amount)
                account_monthly_totals[month] += amount
                grand_monthly_totals[month] += amount
                if current_date.month == 12:
                    current_date = current_date.replace(year=current_date.year + 1, month=1)
                else:
                    current_date = current_date.replace(month=current_date.month + 1)

            closing = get_closing_balance(schedule, end_date)
            account_totals['closing'] += closing
            grand_totals['closing'] += closing

            ws.cell(row=row_num, column=1, value=idx)
            ws.cell(row=row_num, column=2, value=schedule.transaction_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=3, value=schedule.description)
            ws.cell(row=row_num, column=4, value=round(opening, 2))
            ws.cell(row=row_num, column=5, value=round(additions, 2))

            col = 6
            for amount in monthly_amounts:
                ws.cell(row=row_num, column=col, value=round(amount, 2))
                col += 1

            ws.cell(row=row_num, column=col, value=round(closing, 2))
            row_num += 1

        ws.cell(row=row_num, column=1, value="Subtotal")
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
        ws.cell(row=row_num, column=4, value=round(account_totals['opening'], 2))
        ws.cell(row=row_num, column=5, value=round(account_totals['additions'], 2))

        col = 6
        for month in months:
            ws.cell(row=row_num, column=col, value=round(account_monthly_totals[month], 2))
            col += 1

        ws.cell(row=row_num, column=col, value=round(account_totals['closing'], 2))

        for col in range(1, title_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='A5D6A7', end_color='A5D6A7', fill_type='solid')
            cell.alignment = Alignment(horizontal='center' if col <= 3 else 'right', vertical='center')

        row_num += 1

    ws.cell(row=row_num, column=1, value="GRAND TOTAL")
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
    ws.cell(row=row_num, column=4, value=round(grand_totals['opening'], 2))
    ws.cell(row=row_num, column=5, value=round(grand_totals['additions'], 2))

    col = 6
    for month in months:
        ws.cell(row=row_num, column=col, value=round(grand_monthly_totals[month], 2))
        col += 1

    ws.cell(row=row_num, column=col, value=round(grand_totals['closing'], 2))

    for col in range(1, title_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        cell.alignment = Alignment(horizontal='center' if col <= 3 else 'right', vertical='center')

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=2, max_row=row_num, min_col=1, max_col=title_cols):
        for cell in row:
            cell.border = border
            if cell.row > 2:
                if cell.column > 3:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                else:
                    cell.alignment = Alignment(horizontal='left' if cell.column == 3 else 'center', vertical='center')

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


def generate_credit_report_pdf(schedules, start_date, end_date):
    """Generate PDF report with horizontal months grouped by credit account"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16,
                                 textColor=colors.HexColor('#2E7D32'), spaceAfter=15)
    story.append(Paragraph('Monthly Prepayment Report by Credit Account', title_style))

    date_style = ParagraphStyle('DateStyle', parent=styles['Normal'], fontSize=10,
                                textColor=colors.HexColor('#666666'), spaceAfter=12)
    story.append(
        Paragraph(f'Period: {start_date.strftime("%B %d, %Y")} to {end_date.strftime("%B %d, %Y")}', date_style))
    story.append(Spacer(1, 12))

    months = []
    current = start_date
    while current <= end_date:
        months.append(current.strftime('%b'))
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)

    grouped_schedules = get_schedules_by_credit_account(schedules)

    grand_totals = {'opening': 0, 'additions': 0, 'closing': 0}
    grand_monthly_totals = {month: 0 for month in months}

    table_data = []

    headers = ['#', 'Transaction Date', 'Description', 'Opening', 'Period Additions'] + months + ['Closing']
    table_data.append(headers)

    for credit_account, account_schedules in grouped_schedules.items():
        account_header = [credit_account] + [''] * (len(headers) - 1)
        table_data.append(account_header)

        account_totals = {'opening': 0, 'additions': 0, 'closing': 0}
        account_monthly_totals = {month: 0 for month in months}

        for idx, schedule in enumerate(account_schedules, 1):
            opening = get_opening_balance(schedule, start_date)
            account_totals['opening'] += opening
            grand_totals['opening'] += opening

            additions = get_period_additions(schedule, start_date, end_date)
            account_totals['additions'] += additions
            grand_totals['additions'] += additions

            monthly_amounts = []
            current_date = start_date
            for month in months:
                amount = get_monthly_amortization(schedule, current_date.year, current_date.month)
                monthly_amounts.append(amount)
                account_monthly_totals[month] += amount
                grand_monthly_totals[month] += amount
                if current_date.month == 12:
                    current_date = current_date.replace(year=current_date.year + 1, month=1)
                else:
                    current_date = current_date.replace(month=current_date.month + 1)

            closing = get_closing_balance(schedule, end_date)
            account_totals['closing'] += closing
            grand_totals['closing'] += closing

            row = [
                      str(idx),
                      schedule.transaction_date.strftime('%Y-%m-%d'),
                      schedule.description[:25] + '...' if len(schedule.description) > 25 else schedule.description,
                      format_number(opening),
                      format_number(additions)
                  ] + [format_number(amt) for amt in monthly_amounts] + [format_number(closing)]
            table_data.append(row)

        subtotal_row = ['Subtotal', '', '',
                        format_number(account_totals['opening']),
                        format_number(account_totals['additions'])] + \
                       [format_number(account_monthly_totals[month]) for month in months] + \
                       [format_number(account_totals['closing'])]
        table_data.append(subtotal_row)

    grand_total_row = ['GRAND TOTAL', '', '',
                       format_number(grand_totals['opening']),
                       format_number(grand_totals['additions'])] + \
                      [format_number(grand_monthly_totals[month]) for month in months] + \
                      [format_number(grand_totals['closing'])]
    table_data.append(grand_total_row)

    col_widths = [0.4 * inch, 0.9 * inch, 2.0 * inch, 0.9 * inch, 1.0 * inch] + [0.7 * inch] * len(months) + [
        0.9 * inch]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]

    for i, row in enumerate(table_data):
        if row[0] and not row[1] and not row[2] and i > 0:
            style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#388E3C')))
            style.append(('TEXTCOLOR', (0, i), (-1, i), colors.white))
            style.append(('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold'))
            style.append(('FONTSIZE', (0, i), (-1, i), 8))
            style.append(('ALIGN', (0, i), (-1, i), 'LEFT'))
        elif row[0] and row[0] == 'Subtotal':
            style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#A5D6A7')))
            style.append(('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold'))
            style.append(('FONTSIZE', (0, i), (-1, i), 7))
        elif row[0] and row[0] == 'GRAND TOTAL':
            style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FFC000')))
            style.append(('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold'))
            style.append(('FONTSIZE', (0, i), (-1, i), 8))

    for i in range(len(table_data)):
        for j in range(3, len(headers)):
            style.append(('ALIGN', (j, i), (j, i), 'RIGHT'))

    table.setStyle(TableStyle(style))
    story.append(table)

    story.append(Spacer(1, 12))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#666666'))
    story.append(Paragraph(f'Generated on: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}', footer_style))
    story.append(Paragraph(f'Total Records: {len(schedules)}', footer_style))

    doc.build(story)
    buffer.seek(0)
    return buffer


# ============ CREDIT ACCOUNT REPORT ROUTES ============

@app.route('/report/credit', methods=['GET', 'POST'])
def credit_report():
    """Display credit account report form"""
    company_id = session.get('current_company')
    if not company_id:
        company = Company.query.first()
        if company:
            session['current_company'] = company.id
            company_id = company.id

    return render_template('credit_report_form.html')


@app.route('/report/credit/view', methods=['GET', 'POST'])
def credit_report_view():
    """View credit account report in HTML format"""
    company_id = session.get('current_company')
    if not company_id:
        company = Company.query.first()
        if company:
            session['current_company'] = company.id
            company_id = company.id

    if request.method == 'POST':
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
    else:
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

    if not start_date_str or not end_date_str:
        flash('Please select both start and end dates.', 'danger')
        return redirect(url_for('credit_report'))

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        if start_date > end_date:
            flash('Start date must be before end date.', 'danger')
            return redirect(url_for('credit_report'))

        schedules = filter_schedules_for_report(company_id, start_date, end_date)

        if not schedules:
            flash('No schedules found for this period.', 'warning')
            return redirect(url_for('credit_report'))

        months = []
        current = start_date
        while current <= end_date:
            months.append(current.strftime('%b'))
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)

        session['credit_report_start_date'] = start_date_str
        session['credit_report_end_date'] = end_date_str

        grouped_schedules = get_schedules_by_credit_account(schedules)

        report_groups = []
        grand_totals = {'opening': 0, 'additions': 0, 'closing': 0}
        grand_monthly_totals = {month: 0 for month in months}

        for credit_account, account_schedules in grouped_schedules.items():
            account_data = []
            account_totals = {'opening': 0, 'additions': 0, 'closing': 0}
            account_monthly_totals = {month: 0 for month in months}

            for schedule in account_schedules:
                opening = get_opening_balance(schedule, start_date)
                account_totals['opening'] += opening
                grand_totals['opening'] += opening

                additions = get_period_additions(schedule, start_date, end_date)
                account_totals['additions'] += additions
                grand_totals['additions'] += additions

                monthly_amounts = []
                current_date = start_date
                for month in months:
                    amount = get_monthly_amortization(schedule, current_date.year, current_date.month)
                    monthly_amounts.append(amount)
                    account_monthly_totals[month] += amount
                    grand_monthly_totals[month] += amount
                    if current_date.month == 12:
                        current_date = current_date.replace(year=current_date.year + 1, month=1)
                    else:
                        current_date = current_date.replace(month=current_date.month + 1)

                closing = get_closing_balance(schedule, end_date)
                account_totals['closing'] += closing
                grand_totals['closing'] += closing

                account_data.append({
                    'transaction_date': schedule.transaction_date.strftime('%Y-%m-%d'),
                    'description': schedule.description,
                    'opening': opening,
                    'additions': additions,
                    'monthly_amounts': monthly_amounts,
                    'closing': closing
                })

            report_groups.append({
                'credit_account': credit_account,
                'schedules': account_data,
                'totals': account_totals,
                'monthly_totals': account_monthly_totals
            })

        return render_template('credit_report_results.html',
                               report_groups=report_groups,
                               months=months,
                               grand_totals=grand_totals,
                               grand_monthly_totals=grand_monthly_totals,
                               start_date=start_date,
                               end_date=end_date,
                               generated_on=datetime.now(),
                               format_number=format_number)

    except Exception as e:
        flash(f'❌ Error generating report: {str(e)}', 'danger')
        return redirect(url_for('credit_report'))


@app.route('/report/credit/download/pdf')
def download_credit_report_pdf():
    start_date_str = request.args.get('start_date') or session.get('credit_report_start_date')
    end_date_str = request.args.get('end_date') or session.get('credit_report_end_date')

    if not start_date_str or not end_date_str:
        flash('Please generate the report first.', 'warning')
        return redirect(url_for('credit_report'))

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        company_id = session.get('current_company')

        schedules = filter_schedules_for_report(company_id, start_date, end_date)

        if not schedules:
            flash('No schedules found for this period.', 'warning')
            return redirect(url_for('credit_report'))

        pdf_file = generate_credit_report_pdf(schedules, start_date, end_date)
        response = make_response(send_file(pdf_file, mimetype='application/pdf'))
        response.headers['Content-Disposition'] = f'attachment; filename=credit_report_{start_date}_to_{end_date}.pdf'
        return response

    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'danger')
        return redirect(url_for('credit_report'))


@app.route('/report/credit/download/excel')
def download_credit_report_excel():
    start_date_str = request.args.get('start_date') or session.get('credit_report_start_date')
    end_date_str = request.args.get('end_date') or session.get('credit_report_end_date')

    if not start_date_str or not end_date_str:
        flash('Please generate the report first.', 'warning')
        return redirect(url_for('credit_report'))

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        company_id = session.get('current_company')

        schedules = filter_schedules_for_report(company_id, start_date, end_date)

        if not schedules:
            flash('No schedules found for this period.', 'warning')
            return redirect(url_for('credit_report'))

        excel_file = generate_credit_report_excel(schedules, start_date, end_date)
        response = make_response(
            send_file(excel_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
        response.headers['Content-Disposition'] = f'attachment; filename=credit_report_{start_date}_to_{end_date}.xlsx'
        return response

    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'danger')
        return redirect(url_for('credit_report'))


# ============ COMPANY MANAGEMENT ROUTES ============

@app.route('/companies')
def company_list():
    """List all companies"""
    companies = Company.query.order_by(Company.name).all()
    return render_template('company_list.html', companies=companies)


@app.route('/company/create', methods=['GET', 'POST'])
def company_create():
    """Create a new company"""
    if request.method == 'POST':
        try:
            name = request.form.get('name', '').strip()
            code = request.form.get('code', '').strip()
            address = request.form.get('address', '').strip()
            phone = request.form.get('phone', '').strip()
            email = request.form.get('email', '').strip()

            # Validate
            if not name:
                flash('Company name is required.', 'danger')
                return render_template('company_form.html', action='Create')

            if not code:
                flash('Company code is required.', 'danger')
                return render_template('company_form.html', action='Create')

            # Check for duplicate code
            existing = Company.query.filter_by(code=code).first()
            if existing:
                flash(f'Company code "{code}" already exists.', 'danger')
                return render_template('company_form.html', action='Create')

            # Handle logo upload
            logo_filename = None
            if 'logo' in request.files:
                logo_file = request.files['logo']
                if logo_file and logo_file.filename:
                    filename = secure_filename(logo_file.filename)
                    # Create unique filename
                    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                    logo_filename = f"{timestamp}_{filename}"
                    logo_path = os.path.join(app.config['UPLOAD_FOLDER'], logo_filename)
                    logo_file.save(logo_path)

            company = Company(
                name=name,
                code=code,
                address=address,
                phone=phone,
                email=email,
                logo=logo_filename,
                is_active=True
            )
            db.session.add(company)
            db.session.commit()

            flash(f'✅ Company "{name}" created successfully!', 'success')
            return redirect(url_for('company_list'))

        except Exception as e:
            flash(f'❌ Error creating company: {str(e)}', 'danger')
            db.session.rollback()

    return render_template('company_form.html', action='Create')


@app.route('/company/<int:pk>/edit', methods=['GET', 'POST'])
def company_edit(pk):
    """Edit a company"""
    company = Company.query.get_or_404(pk)

    if request.method == 'POST':
        try:
            company.name = request.form.get('name', '').strip()
            company.code = request.form.get('code', '').strip()
            company.address = request.form.get('address', '').strip()
            company.phone = request.form.get('phone', '').strip()
            company.email = request.form.get('email', '').strip()
            company.is_active = request.form.get('is_active') == 'on'

            # Validate
            if not company.name:
                flash('Company name is required.', 'danger')
                return render_template('company_form.html', action='Edit', company=company)

            if not company.code:
                flash('Company code is required.', 'danger')
                return render_template('company_form.html', action='Edit', company=company)

            # Check for duplicate code (excluding current company)
            existing = Company.query.filter(Company.code == company.code, Company.id != pk).first()
            if existing:
                flash(f'Company code "{company.code}" already exists.', 'danger')
                return render_template('company_form.html', action='Edit', company=company)

            # Handle logo upload
            if 'logo' in request.files:
                logo_file = request.files['logo']
                if logo_file and logo_file.filename:
                    # Delete old logo if exists
                    if company.logo:
                        old_logo_path = os.path.join(app.config['UPLOAD_FOLDER'], company.logo)
                        if os.path.exists(old_logo_path):
                            os.remove(old_logo_path)

                    filename = secure_filename(logo_file.filename)
                    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                    logo_filename = f"{timestamp}_{filename}"
                    logo_path = os.path.join(app.config['UPLOAD_FOLDER'], logo_filename)
                    logo_file.save(logo_path)
                    company.logo = logo_filename

            db.session.commit()
            flash(f'✅ Company "{company.name}" updated successfully!', 'success')
            return redirect(url_for('company_list'))

        except Exception as e:
            flash(f'❌ Error updating company: {str(e)}', 'danger')
            db.session.rollback()

    return render_template('company_form.html', action='Edit', company=company)


@app.route('/company/<int:pk>/delete', methods=['POST'])
def company_delete(pk):
    """Delete a company"""
    company = Company.query.get_or_404(pk)

    # Check if company has schedules
    schedules = PrepaymentSchedule.query.filter_by(company_id=pk).count()
    if schedules > 0:
        flash(
            f'❌ Cannot delete company "{company.name}" because it has {schedules} schedules. Delete the schedules first.',
            'danger')
        return redirect(url_for('company_list'))

    try:
        # Delete logo file if exists
        if company.logo:
            logo_path = os.path.join(app.config['UPLOAD_FOLDER'], company.logo)
            if os.path.exists(logo_path):
                os.remove(logo_path)

        db.session.delete(company)
        db.session.commit()
        flash(f'✅ Company "{company.name}" deleted successfully!', 'success')
    except Exception as e:
        flash(f'❌ Error deleting company: {str(e)}', 'danger')
        db.session.rollback()

    return redirect(url_for('company_list'))


# ============ COMBINED DEBIT & CREDIT REPORT FUNCTIONS ============

def get_schedules_grouped(schedules):
    """Group schedules by both debit and credit account"""
    grouped = {}
    for schedule in schedules:
        key = f"{schedule.debit_account} | {schedule.credit_account}"
        if key not in grouped:
            grouped[key] = {
                'debit_account': schedule.debit_account,
                'credit_account': schedule.credit_account,
                'schedules': []
            }
        grouped[key]['schedules'].append(schedule)
    return grouped


def generate_combined_report_excel(schedules, start_date, end_date):
    """Generate Excel report with both debit and credit accounts"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Combined Report"

    months = []
    current = start_date
    while current <= end_date:
        months.append(current.strftime('%b'))
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)

    grouped_schedules = get_schedules_grouped(schedules)

    title_cols = len(months) + 7  # #, Debit, Credit, Description, Opening, Additions, months, Closing
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_cols)
    title_cell = ws.cell(row=1, column=1,
                         value=f"Combined Prepayment Report: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    headers = ['#', 'Debit Account', 'Credit Account', 'Description', 'Opening', 'Period Additions'] + months + ['Closing']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='6C3483', end_color='6C3483', fill_type='solid')  # Purple theme
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = 18 if col <= 4 else 15

    row_num = 3
    grand_totals = {'opening': 0, 'additions': 0, 'closing': 0}
    grand_monthly_totals = {month: 0 for month in months}

    for key, group in grouped_schedules.items():
        # Account header row
        account_header_cell = ws.cell(row=row_num, column=1,
                                      value=f"DEBIT: {group['debit_account']} | CREDIT: {group['credit_account']}")
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=title_cols)
        account_header_cell.font = Font(bold=True, size=12, color='FFFFFF')
        account_header_cell.fill = PatternFill(start_color='7D3C98', end_color='7D3C98', fill_type='solid')
        account_header_cell.alignment = Alignment(horizontal='left', vertical='center')
        row_num += 1

        account_totals = {'opening': 0, 'additions': 0, 'closing': 0}
        account_monthly_totals = {month: 0 for month in months}

        for idx, schedule in enumerate(group['schedules'], 1):
            opening = get_opening_balance(schedule, start_date)
            account_totals['opening'] += opening
            grand_totals['opening'] += opening

            additions = get_period_additions(schedule, start_date, end_date)
            account_totals['additions'] += additions
            grand_totals['additions'] += additions

            monthly_amounts = []
            current_date = start_date
            for month in months:
                amount = get_monthly_amortization(schedule, current_date.year, current_date.month)
                monthly_amounts.append(amount)
                account_monthly_totals[month] += amount
                grand_monthly_totals[month] += amount
                if current_date.month == 12:
                    current_date = current_date.replace(year=current_date.year + 1, month=1)
                else:
                    current_date = current_date.replace(month=current_date.month + 1)

            closing = get_closing_balance(schedule, end_date)
            account_totals['closing'] += closing
            grand_totals['closing'] += closing

            ws.cell(row=row_num, column=1, value=idx)
            ws.cell(row=row_num, column=2, value=schedule.debit_account)
            ws.cell(row=row_num, column=3, value=schedule.credit_account)
            ws.cell(row=row_num, column=4, value=schedule.description)
            ws.cell(row=row_num, column=5, value=round(opening, 2))
            ws.cell(row=row_num, column=6, value=round(additions, 2))

            col = 7
            for amount in monthly_amounts:
                ws.cell(row=row_num, column=col, value=round(amount, 2))
                col += 1

            ws.cell(row=row_num, column=col, value=round(closing, 2))
            row_num += 1

        # Subtotal row
        ws.cell(row=row_num, column=1, value="Subtotal")
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
        ws.cell(row=row_num, column=5, value=round(account_totals['opening'], 2))
        ws.cell(row=row_num, column=6, value=round(account_totals['additions'], 2))

        col = 7
        for month in months:
            ws.cell(row=row_num, column=col, value=round(account_monthly_totals[month], 2))
            col += 1

        ws.cell(row=row_num, column=col, value=round(account_totals['closing'], 2))

        for col in range(1, title_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D2B4DE', end_color='D2B4DE', fill_type='solid')
            cell.alignment = Alignment(horizontal='center' if col <= 4 else 'right', vertical='center')

        row_num += 1

    # Grand total row
    ws.cell(row=row_num, column=1, value="GRAND TOTAL")
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
    ws.cell(row=row_num, column=5, value=round(grand_totals['opening'], 2))
    ws.cell(row=row_num, column=6, value=round(grand_totals['additions'], 2))

    col = 7
    for month in months:
        ws.cell(row=row_num, column=col, value=round(grand_monthly_totals[month], 2))
        col += 1

    ws.cell(row=row_num, column=col, value=round(grand_totals['closing'], 2))

    for col in range(1, title_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        cell.alignment = Alignment(horizontal='center' if col <= 4 else 'right', vertical='center')

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=2, max_row=row_num, min_col=1, max_col=title_cols):
        for cell in row:
            cell.border = border
            if cell.row > 2:
                if cell.column > 4:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                else:
                    cell.alignment = Alignment(horizontal='left' if cell.column <= 4 else 'center', vertical='center')

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


def generate_combined_report_pdf(schedules, start_date, end_date):
    """Generate PDF report with both debit and credit accounts"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16,
                                 textColor=colors.HexColor('#6C3483'), spaceAfter=15)
    story.append(Paragraph('Combined Prepayment Report', title_style))

    date_style = ParagraphStyle('DateStyle', parent=styles['Normal'], fontSize=10,
                                textColor=colors.HexColor('#666666'), spaceAfter=12)
    story.append(
        Paragraph(f'Period: {start_date.strftime("%B %d, %Y")} to {end_date.strftime("%B %d, %Y")}', date_style))
    story.append(Spacer(1, 12))

    months = []
    current = start_date
    while current <= end_date:
        months.append(current.strftime('%b'))
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)

    grouped_schedules = get_schedules_grouped(schedules)

    grand_totals = {'opening': 0, 'additions': 0, 'closing': 0}
    grand_monthly_totals = {month: 0 for month in months}

    table_data = []

    headers = ['#', 'Debit Account', 'Credit Account', 'Description', 'Opening', 'Period Additions'] + months + ['Closing']
    table_data.append(headers)

    for key, group in grouped_schedules.items():
        account_header = [f'DEBIT: {group["debit_account"]} | CREDIT: {group["credit_account"]}'] + [''] * (len(headers) - 1)
        table_data.append(account_header)

        account_totals = {'opening': 0, 'additions': 0, 'closing': 0}
        account_monthly_totals = {month: 0 for month in months}

        for idx, schedule in enumerate(group['schedules'], 1):
            opening = get_opening_balance(schedule, start_date)
            account_totals['opening'] += opening
            grand_totals['opening'] += opening

            additions = get_period_additions(schedule, start_date, end_date)
            account_totals['additions'] += additions
            grand_totals['additions'] += additions

            monthly_amounts = []
            current_date = start_date
            for month in months:
                amount = get_monthly_amortization(schedule, current_date.year, current_date.month)
                monthly_amounts.append(amount)
                account_monthly_totals[month] += amount
                grand_monthly_totals[month] += amount
                if current_date.month == 12:
                    current_date = current_date.replace(year=current_date.year + 1, month=1)
                else:
                    current_date = current_date.replace(month=current_date.month + 1)

            closing = get_closing_balance(schedule, end_date)
            account_totals['closing'] += closing
            grand_totals['closing'] += closing

            row = [
                str(idx),
                schedule.debit_account,
                schedule.credit_account,
                schedule.description[:25] + '...' if len(schedule.description) > 25 else schedule.description,
                format_number(opening),
                format_number(additions)
            ] + [format_number(amt) for amt in monthly_amounts] + [format_number(closing)]
            table_data.append(row)

        subtotal_row = ['Subtotal', '', '', '',
                        format_number(account_totals['opening']),
                        format_number(account_totals['additions'])] + \
                       [format_number(account_monthly_totals[month]) for month in months] + \
                       [format_number(account_totals['closing'])]
        table_data.append(subtotal_row)

    grand_total_row = ['GRAND TOTAL', '', '', '',
                       format_number(grand_totals['opening']),
                       format_number(grand_totals['additions'])] + \
                      [format_number(grand_monthly_totals[month]) for month in months] + \
                      [format_number(grand_totals['closing'])]
    table_data.append(grand_total_row)

    col_widths = [0.4 * inch, 0.9 * inch, 0.9 * inch, 1.8 * inch, 0.8 * inch, 0.8 * inch] + [0.7 * inch] * len(months) + [0.8 * inch]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C3483')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]

    for i, row in enumerate(table_data):
        if row[0] and row[0].startswith('DEBIT:'):
            style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#7D3C98')))
            style.append(('TEXTCOLOR', (0, i), (-1, i), colors.white))
            style.append(('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold'))
            style.append(('FONTSIZE', (0, i), (-1, i), 8))
            style.append(('ALIGN', (0, i), (-1, i), 'LEFT'))
        elif row[0] and row[0] == 'Subtotal':
            style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#D2B4DE')))
            style.append(('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold'))
            style.append(('FONTSIZE', (0, i), (-1, i), 7))
        elif row[0] and row[0] == 'GRAND TOTAL':
            style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FFC000')))
            style.append(('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold'))
            style.append(('FONTSIZE', (0, i), (-1, i), 8))

    for i in range(len(table_data)):
        for j in range(4, len(headers)):
            style.append(('ALIGN', (j, i), (j, i), 'RIGHT'))

    table.setStyle(TableStyle(style))
    story.append(table)

    story.append(Spacer(1, 12))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#666666'))
    story.append(Paragraph(f'Generated on: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}', footer_style))
    story.append(Paragraph(f'Total Records: {len(schedules)}', footer_style))

    doc.build(story)
    buffer.seek(0)
    return buffer


# ============ COMBINED REPORT ROUTES ============

@app.route('/report/combined', methods=['GET', 'POST'])
def combined_report():
    """Display combined report form"""
    company_id = session.get('current_company')
    if not company_id:
        company = Company.query.first()
        if company:
            session['current_company'] = company.id
            company_id = company.id

    return render_template('combined_report_form.html')


@app.route('/report/combined/view', methods=['GET', 'POST'])
def combined_report_view():
    """View combined report in HTML format"""
    company_id = session.get('current_company')
    if not company_id:
        company = Company.query.first()
        if company:
            session['current_company'] = company.id
            company_id = company.id

    if request.method == 'POST':
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
    else:
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

    if not start_date_str or not end_date_str:
        flash('Please select both start and end dates.', 'danger')
        return redirect(url_for('combined_report'))

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        if start_date > end_date:
            flash('Start date must be before end date.', 'danger')
            return redirect(url_for('combined_report'))

        schedules = filter_schedules_for_report(company_id, start_date, end_date)

        if not schedules:
            flash('No schedules found for this period.', 'warning')
            return redirect(url_for('combined_report'))

        months = []
        current = start_date
        while current <= end_date:
            months.append(current.strftime('%b'))
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)

        session['combined_report_start_date'] = start_date_str
        session['combined_report_end_date'] = end_date_str

        grouped_schedules = get_schedules_grouped(schedules)

        report_groups = []
        grand_totals = {'opening': 0, 'additions': 0, 'closing': 0}
        grand_monthly_totals = {month: 0 for month in months}

        for key, group in grouped_schedules.items():
            account_data = []
            account_totals = {'opening': 0, 'additions': 0, 'closing': 0}
            account_monthly_totals = {month: 0 for month in months}

            for schedule in group['schedules']:
                opening = get_opening_balance(schedule, start_date)
                account_totals['opening'] += opening
                grand_totals['opening'] += opening

                additions = get_period_additions(schedule, start_date, end_date)
                account_totals['additions'] += additions
                grand_totals['additions'] += additions

                monthly_amounts = []
                current_date = start_date
                for month in months:
                    amount = get_monthly_amortization(schedule, current_date.year, current_date.month)
                    monthly_amounts.append(amount)
                    account_monthly_totals[month] += amount
                    grand_monthly_totals[month] += amount
                    if current_date.month == 12:
                        current_date = current_date.replace(year=current_date.year + 1, month=1)
                    else:
                        current_date = current_date.replace(month=current_date.month + 1)

                closing = get_closing_balance(schedule, end_date)
                account_totals['closing'] += closing
                grand_totals['closing'] += closing

                account_data.append({
                    'transaction_date': schedule.transaction_date.strftime('%Y-%m-%d'),
                    'debit_account': schedule.debit_account,
                    'credit_account': schedule.credit_account,
                    'description': schedule.description,
                    'opening': opening,
                    'additions': additions,
                    'monthly_amounts': monthly_amounts,
                    'closing': closing
                })

            report_groups.append({
                'key': key,
                'debit_account': group['debit_account'],
                'credit_account': group['credit_account'],
                'schedules': account_data,
                'totals': account_totals,
                'monthly_totals': account_monthly_totals
            })

        return render_template('combined_report_results.html',
                               report_groups=report_groups,
                               months=months,
                               grand_totals=grand_totals,
                               grand_monthly_totals=grand_monthly_totals,
                               start_date=start_date,
                               end_date=end_date,
                               generated_on=datetime.now(),
                               format_number=format_number)

    except Exception as e:
        flash(f'❌ Error generating report: {str(e)}', 'danger')
        return redirect(url_for('combined_report'))


@app.route('/report/combined/download/pdf')
def download_combined_report_pdf():
    start_date_str = request.args.get('start_date') or session.get('combined_report_start_date')
    end_date_str = request.args.get('end_date') or session.get('combined_report_end_date')

    if not start_date_str or not end_date_str:
        flash('Please generate the report first.', 'warning')
        return redirect(url_for('combined_report'))

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        company_id = session.get('current_company')

        schedules = filter_schedules_for_report(company_id, start_date, end_date)

        if not schedules:
            flash('No schedules found for this period.', 'warning')
            return redirect(url_for('combined_report'))

        pdf_file = generate_combined_report_pdf(schedules, start_date, end_date)
        response = make_response(send_file(pdf_file, mimetype='application/pdf'))
        response.headers['Content-Disposition'] = f'attachment; filename=combined_report_{start_date}_to_{end_date}.pdf'
        return response

    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'danger')
        return redirect(url_for('combined_report'))


@app.route('/report/combined/download/excel')
def download_combined_report_excel():
    start_date_str = request.args.get('start_date') or session.get('combined_report_start_date')
    end_date_str = request.args.get('end_date') or session.get('combined_report_end_date')

    if not start_date_str or not end_date_str:
        flash('Please generate the report first.', 'warning')
        return redirect(url_for('combined_report'))

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        company_id = session.get('current_company')

        schedules = filter_schedules_for_report(company_id, start_date, end_date)

        if not schedules:
            flash('No schedules found for this period.', 'warning')
            return redirect(url_for('combined_report'))

        excel_file = generate_combined_report_excel(schedules, start_date, end_date)
        response = make_response(
            send_file(excel_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
        response.headers['Content-Disposition'] = f'attachment; filename=combined_report_{start_date}_to_{end_date}.xlsx'
        return response

    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'danger')
        return redirect(url_for('combined_report'))

@app.route('/switch-company', methods=['GET', 'POST'])
def switch_company():
    companies = Company.query.filter_by(is_active=True).all()

    if not companies:
        flash('No companies available. Creating default...', 'info')
        setup_default_company()
        companies = Company.query.filter_by(is_active=True).all()

    if request.method == 'POST':
        company_id = request.form.get('company')
        if company_id:
            company = Company.query.get(int(company_id))
            if company:
                session['current_company'] = int(company_id)
                flash(f'✅ Switched to {company.name}', 'success')
            else:
                flash('❌ Company not found.', 'danger')
        return redirect(url_for('dashboard'))

    current_company_id = session.get('current_company')
    return render_template('company_switch.html',
                           companies=companies,
                           current_company_id=current_company_id)


# ============ RUN APP ============
if __name__ == '__main__':
    # Only use debug mode locally
    debug_mode = not os.environ.get('RENDER')

    print("\n" + "=" * 50)
    print("🚀 Prepayment Scheduler is running!")
    print("📍 http://localhost:5000")
    print("=" * 50 + "\n")

    app.run(debug=debug_mode, host='0.0.0.0', port=5000)