from flask import Flask, render_template, request, jsonify, session, send_file, redirect, url_for, flash
from functools import wraps
from dataclasses import dataclass, field
from typing import List, Dict
import json
import os
from datetime import datetime
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from werkzeug.security import generate_password_hash, check_password_hash
import uuid
import traceback
import logging
from sqlalchemy import create_engine, Column, String, Float, Integer, Boolean, DateTime, Text, JSON, ForeignKey
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, relationship
from sqlalchemy.pool import NullPool

# Try to import psycopg2 for PostgreSQL support
try:
    import psycopg2
    print("✅ psycopg2 imported successfully")
except ImportError:
    print("ℹ️ psycopg2 not installed - using SQLite (local development)")
    psycopg2 = None

logging.basicConfig(level=logging.DEBUG)

# Database configuration
IS_RENDER = os.environ.get('RENDER', False)
DATABASE_URL = os.environ.get('DATABASE_URL', '')

# Get the base path for routing - this is set by the wrapper
BASE_PATH = os.environ.get('BASE_PATH', '')
if BASE_PATH:
    print(f"📁 App mounted at: {BASE_PATH}")

# SQLAlchemy setup
Base = declarative_base()


# Define database models
class User(Base):
    __tablename__ = 'users'

    id = Column(String(36), primary_key=True)
    username = Column(String(80), unique=True, nullable=False)
    password_hash = Column(String(200), nullable=False)
    email = Column(String(120))
    full_name = Column(String(120))
    role = Column(String(50), default='user')
    created_date = Column(DateTime, default=datetime.utcnow)
    last_login = Column(DateTime, nullable=True)
    is_active = Column(Boolean, default=True)


class Project(Base):
    __tablename__ = 'projects'

    id = Column(String(36), primary_key=True)
    project_name = Column(String(200), nullable=False)
    user_id = Column(String(36), ForeignKey('users.id'))
    created_date = Column(DateTime, default=datetime.utcnow)
    modified_date = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    implementation_items = Column(JSON, default=[])
    logistics_items = Column(JSON, default=[])
    software_items = Column(JSON, default=[])

    annual_maintenance_cost = Column(Float, default=0.0)
    expected_annual_benefit = Column(Float, default=0.0)
    project_lifetime_years = Column(Float, default=3.0)
    discount_rate = Column(Float, default=10.0)
    markup_percentages = Column(JSON, default=lambda: {
        "implementation": 0, "software": 0, "logistics": 0, "maintenance": 0
    })

    currency_code = Column(String(10), default='USD')
    currency_symbol = Column(String(10), default='$')
    exchange_rate = Column(Float, default=1.0)

    approval_status = Column(String(20), default='pending')
    approved_by = Column(String(80), nullable=True)
    approved_date = Column(DateTime, nullable=True)
    approval_comments = Column(Text, default='')

    user = relationship("User")


class UserProjectAssignment(Base):
    __tablename__ = 'user_project_assignments'

    id = Column(Integer, primary_key=True, autoincrement=True)
    project_id = Column(String(36), ForeignKey('projects.id'))
    user_id = Column(String(36), ForeignKey('users.id'))


class Setting(Base):
    __tablename__ = 'settings'

    id = Column(Integer, primary_key=True, autoincrement=True)
    key = Column(String(100), unique=True, nullable=False)
    value = Column(Text)


class CurrencyRate(Base):
    __tablename__ = 'currency_rates'

    id = Column(Integer, primary_key=True, autoincrement=True)
    currency_code = Column(String(10), unique=True, nullable=False)
    rate = Column(Float, default=1.0)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


def get_engine():
    """Create database engine based on environment"""
    if IS_RENDER and DATABASE_URL and psycopg2 is not None:
        try:
            print(f"🔗 Connecting to PostgreSQL on Render...")
            return create_engine(DATABASE_URL, poolclass=NullPool, pool_pre_ping=True)
        except Exception as e:
            print(f"❌ Error creating PostgreSQL engine: {e}")
            print("⚠️ Falling back to SQLite...")
            return create_sqlite_engine()
    else:
        return create_sqlite_engine()


def create_sqlite_engine():
    """Create SQLite engine for local development"""
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'project_costing.db')
    print(f"📁 Using SQLite database: {db_path}")
    return create_engine(f'sqlite:///{db_path}', connect_args={'check_same_thread': False})


# Create engine and session
engine = get_engine()
SessionLocal = sessionmaker(bind=engine)


def init_db():
    """Initialize database tables"""
    print("🔄 Initializing database...")
    Base.metadata.create_all(engine)

    db_session = SessionLocal()
    try:
        # Check if settings exist
        settings_count = db_session.query(Setting).count()
        if settings_count == 0:
            print("📝 Creating default settings...")
            default_settings = {
                'company_name': 'Your Company Name',
                'company_address': '123 Business Street, City, State 12345',
                'company_phone': '+1 (555) 123-4567',
                'company_email': 'info@yourcompany.com',
                'company_website': 'www.yourcompany.com',
                'system_name': 'Project Costing System',
                'report_footer': 'This report is confidential and generated by the Project Costing System.',
                'logo_filename': '',
                'default_currency': 'USD',
                'base_currency': 'USD'
            }
            for key, value in default_settings.items():
                setting = Setting(key=key, value=json.dumps(value))
                db_session.add(setting)
            db_session.commit()

        # Check if default users exist
        admin = db_session.query(User).filter(User.username == 'admin').first()
        if not admin:
            print("👤 Creating default users...")
            default_users = [
                ('admin', 'admin123', 'admin@example.com', 'Administrator', 'admin'),
                ('viewer', 'viewer123', 'viewer@example.com', 'Viewer User', 'viewer'),
                ('technical_lead', 'tech123', 'tech@example.com', 'Technical Lead', 'technical_lead'),
                ('procurement', 'proc123', 'proc@example.com', 'Procurement Manager', 'procurement'),
                ('ceo', 'ceo123', 'ceo@example.com', 'CEO', 'ceo')
            ]

            for username, password, email, full_name, role in default_users:
                user = User(
                    id=str(uuid.uuid4()),
                    username=username,
                    password_hash=generate_password_hash(password),
                    email=email,
                    full_name=full_name,
                    role=role,
                    is_active=True
                )
                db_session.add(user)
            db_session.commit()

            print("=" * 50)
            print("✅ Default users created:")
            print("  Admin: admin / admin123")
            print("  Viewer: viewer / viewer123")
            print("  Technical Lead: technical_lead / tech123")
            print("  Procurement: procurement / proc123")
            print("  CEO: ceo / ceo123")
            print("=" * 50)

        # Check if currency rates exist
        rates_count = db_session.query(CurrencyRate).count()
        if rates_count == 0:
            print("💰 Creating default currency rates...")
            default_rates = {
                'USD': 1.0, 'EUR': 0.85, 'GBP': 0.73, 'JPY': 110.0,
                'CNY': 6.45, 'INR': 74.0, 'AUD': 1.35, 'CAD': 1.25,
                'CHF': 0.92, 'SGD': 1.35, 'HKD': 7.78, 'NZD': 1.43,
                'KRW': 1150.0, 'RUB': 73.5, 'BRL': 5.25, 'ZAR': 14.5,
                'AED': 3.67, 'SAR': 3.75, 'MYR': 4.15, 'THB': 33.0,
                'VND': 23000, 'IDR': 14400, 'PHP': 50.0, 'PKR': 160.0,
                'BDT': 85.0, 'NGN': 410.0, 'EGP': 15.7, 'TRY': 8.5,
                'MXN': 20.0, 'SEK': 8.5, 'NOK': 8.6, 'DKK': 6.3,
                'PLN': 3.8, 'CZK': 21.5, 'HUF': 300.0, 'ILS': 3.2,
                'CLP': 780.0, 'PEN': 3.8, 'COP': 3800.0
            }
            for code, rate in default_rates.items():
                currency_rate = CurrencyRate(currency_code=code, rate=rate)
                db_session.add(currency_rate)
            db_session.commit()

        print("✅ Database initialization complete!")

    except Exception as e:
        print(f"❌ Error initializing database: {e}")
        db_session.rollback()
    finally:
        db_session.close()


# Initialize database
init_db()

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'svg'}
UPLOAD_FOLDER = 'static'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'your-secret-key-here-change-in-production-2024')

EXPORT_FOLDER = 'exports'
REPORT_FOLDER = 'reports'
os.makedirs(EXPORT_FOLDER, exist_ok=True)
os.makedirs(REPORT_FOLDER, exist_ok=True)

# ===================== CURRENCY DEFINITIONS =====================
SUPPORTED_CURRENCIES = {
    'USD': {'symbol': '$', 'name': 'US Dollar', 'code': 'USD', 'decimal_places': 2},
    'EUR': {'symbol': '€', 'name': 'Euro', 'code': 'EUR', 'decimal_places': 2},
    'GBP': {'symbol': '£', 'name': 'British Pound', 'code': 'GBP', 'decimal_places': 2},
    'JPY': {'symbol': '¥', 'name': 'Japanese Yen', 'code': 'JPY', 'decimal_places': 0},
    'CNY': {'symbol': '¥', 'name': 'Chinese Yuan', 'code': 'CNY', 'decimal_places': 2},
    'INR': {'symbol': '₹', 'name': 'Indian Rupee', 'code': 'INR', 'decimal_places': 2},
    'AUD': {'symbol': 'A$', 'name': 'Australian Dollar', 'code': 'AUD', 'decimal_places': 2},
    'CAD': {'symbol': 'C$', 'name': 'Canadian Dollar', 'code': 'CAD', 'decimal_places': 2},
    'CHF': {'symbol': 'Fr', 'name': 'Swiss Franc', 'code': 'CHF', 'decimal_places': 2},
    'SGD': {'symbol': 'S$', 'name': 'Singapore Dollar', 'code': 'SGD', 'decimal_places': 2},
    'HKD': {'symbol': 'HK$', 'name': 'Hong Kong Dollar', 'code': 'HKD', 'decimal_places': 2},
    'NZD': {'symbol': 'NZ$', 'name': 'New Zealand Dollar', 'code': 'NZD', 'decimal_places': 2},
    'KRW': {'symbol': '₩', 'name': 'South Korean Won', 'code': 'KRW', 'decimal_places': 0},
    'RUB': {'symbol': '₽', 'name': 'Russian Ruble', 'code': 'RUB', 'decimal_places': 2},
    'BRL': {'symbol': 'R$', 'name': 'Brazilian Real', 'code': 'BRL', 'decimal_places': 2},
    'ZAR': {'symbol': 'R', 'name': 'South African Rand', 'code': 'ZAR', 'decimal_places': 2},
    'AED': {'symbol': 'د.إ', 'name': 'UAE Dirham', 'code': 'AED', 'decimal_places': 2},
    'SAR': {'symbol': '﷼', 'name': 'Saudi Riyal', 'code': 'SAR', 'decimal_places': 2},
    'MYR': {'symbol': 'RM', 'name': 'Malaysian Ringgit', 'code': 'MYR', 'decimal_places': 2},
    'THB': {'symbol': '฿', 'name': 'Thai Baht', 'code': 'THB', 'decimal_places': 2},
    'VND': {'symbol': '₫', 'name': 'Vietnamese Dong', 'code': 'VND', 'decimal_places': 0},
    'IDR': {'symbol': 'Rp', 'name': 'Indonesian Rupiah', 'code': 'IDR', 'decimal_places': 0},
    'PHP': {'symbol': '₱', 'name': 'Philippine Peso', 'code': 'PHP', 'decimal_places': 2},
    'PKR': {'symbol': '₨', 'name': 'Pakistani Rupee', 'code': 'PKR', 'decimal_places': 2},
    'BDT': {'symbol': '৳', 'name': 'Bangladeshi Taka', 'code': 'BDT', 'decimal_places': 2},
    'NGN': {'symbol': '₦', 'name': 'Nigerian Naira', 'code': 'NGN', 'decimal_places': 2},
    'EGP': {'symbol': '£', 'name': 'Egyptian Pound', 'code': 'EGP', 'decimal_places': 2},
    'TRY': {'symbol': '₺', 'name': 'Turkish Lira', 'code': 'TRY', 'decimal_places': 2},
    'MXN': {'symbol': '$', 'name': 'Mexican Peso', 'code': 'MXN', 'decimal_places': 2},
    'SEK': {'symbol': 'kr', 'name': 'Swedish Krona', 'code': 'SEK', 'decimal_places': 2},
    'NOK': {'symbol': 'kr', 'name': 'Norwegian Krone', 'code': 'NOK', 'decimal_places': 2},
    'DKK': {'symbol': 'kr', 'name': 'Danish Krone', 'code': 'DKK', 'decimal_places': 2},
    'PLN': {'symbol': 'zł', 'name': 'Polish Zloty', 'code': 'PLN', 'decimal_places': 2},
    'CZK': {'symbol': 'Kč', 'name': 'Czech Koruna', 'code': 'CZK', 'decimal_places': 2},
    'HUF': {'symbol': 'Ft', 'name': 'Hungarian Forint', 'code': 'HUF', 'decimal_places': 0},
    'ILS': {'symbol': '₪', 'name': 'Israeli Shekel', 'code': 'ILS', 'decimal_places': 2},
    'CLP': {'symbol': '$', 'name': 'Chilean Peso', 'code': 'CLP', 'decimal_places': 0},
    'PEN': {'symbol': 'S/', 'name': 'Peruvian Sol', 'code': 'PEN', 'decimal_places': 2},
    'COP': {'symbol': '$', 'name': 'Colombian Peso', 'code': 'COP', 'decimal_places': 0},
    'GHS': {'symbol': '₵', 'name': 'Ghanaian Cedi', 'code': 'GHS', 'decimal_places': 2}
}

def get_default_exchange_rates():
    """Return default exchange rates (base: USD)"""
    return {
        'USD': 1.0, 'EUR': 0.85, 'GBP': 0.73, 'JPY': 110.0,
        'CNY': 6.45, 'INR': 74.0, 'AUD': 1.35, 'CAD': 1.25,
        'CHF': 0.92, 'SGD': 1.35, 'HKD': 7.78, 'NZD': 1.43,
        'KRW': 1150.0, 'RUB': 73.5, 'BRL': 5.25, 'ZAR': 14.5,
        'AED': 3.67, 'SAR': 3.75, 'MYR': 4.15, 'THB': 33.0,
        'VND': 23000, 'IDR': 14400, 'PHP': 50.0, 'PKR': 160.0,
        'BDT': 85.0, 'NGN': 410.0, 'EGP': 15.7, 'TRY': 8.5,
        'MXN': 20.0, 'SEK': 8.5, 'NOK': 8.6, 'DKK': 6.3,
        'PLN': 3.8, 'CZK': 21.5, 'HUF': 300.0, 'ILS': 3.2,
        'CLP': 780.0, 'PEN': 3.8, 'COP': 3800.0
    }

# ===================== DATABASE HELPER FUNCTIONS =====================

def get_db():
    db = SessionLocal()
    try:
        return db
    except:
        db.close()
        raise


def load_currency_rates():
    db = get_db()
    try:
        rates = {}
        currency_rates = db.query(CurrencyRate).all()
        for cr in currency_rates:
            rates[cr.currency_code] = cr.rate

        # If no rates found, return empty dict (will use defaults in route)
        return rates
    finally:
        db.close()

def save_currency_rates(rates):
    db = get_db()
    try:
        for code, rate in rates.items():
            currency_rate = db.query(CurrencyRate).filter(CurrencyRate.currency_code == code).first()
            if currency_rate:
                currency_rate.rate = rate
                currency_rate.updated_at = datetime.utcnow()
            else:
                currency_rate = CurrencyRate(currency_code=code, rate=rate)
                db.add(currency_rate)
        db.commit()
    except Exception as e:
        db.rollback()
        raise e
    finally:
        db.close()


def load_settings():
    db = get_db()
    try:
        settings = {}
        all_settings = db.query(Setting).all()
        for setting in all_settings:
            try:
                settings[setting.key] = json.loads(setting.value)
            except:
                settings[setting.key] = setting.value
        return settings
    finally:
        db.close()


def save_settings(settings_dict):
    db = get_db()
    try:
        for key, value in settings_dict.items():
            setting = db.query(Setting).filter(Setting.key == key).first()
            value_json = json.dumps(value)
            if setting:
                setting.value = value_json
            else:
                setting = Setting(key=key, value=value_json)
                db.add(setting)
        db.commit()
    except Exception as e:
        db.rollback()
        raise e
    finally:
        db.close()


def load_users():
    db = get_db()
    try:
        users = db.query(User).all()
        return [{
            'id': u.id,
            'username': u.username,
            'password_hash': u.password_hash,
            'email': u.email,
            'full_name': u.full_name,
            'role': u.role,
            'created_date': u.created_date.isoformat() if u.created_date else None,
            'last_login': u.last_login.isoformat() if u.last_login else None,
            'is_active': u.is_active
        } for u in users]
    finally:
        db.close()


def get_user_by_username(username):
    db = get_db()
    try:
        user = db.query(User).filter(User.username == username).first()
        if user:
            return {
                'id': user.id,
                'username': user.username,
                'password_hash': user.password_hash,
                'email': user.email,
                'full_name': user.full_name,
                'role': user.role,
                'created_date': user.created_date.isoformat() if user.created_date else None,
                'last_login': user.last_login.isoformat() if user.last_login else None,
                'is_active': user.is_active
            }
        return None
    finally:
        db.close()


def get_user_by_id(user_id):
    db = get_db()
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if user:
            return {
                'id': user.id,
                'username': user.username,
                'password_hash': user.password_hash,
                'email': user.email,
                'full_name': user.full_name,
                'role': user.role,
                'created_date': user.created_date.isoformat() if user.created_date else None,
                'last_login': user.last_login.isoformat() if user.last_login else None,
                'is_active': user.is_active
            }
        return None
    finally:
        db.close()


def create_user(username, password, email="", full_name="", role="user"):
    db = get_db()
    try:
        if db.query(User).filter(User.username == username).first():
            return None

        new_user = User(
            id=str(uuid.uuid4()),
            username=username,
            password_hash=generate_password_hash(password),
            email=email,
            full_name=full_name,
            role=role,
            is_active=True
        )
        db.add(new_user)
        db.commit()

        return {
            'id': new_user.id,
            'username': new_user.username,
            'email': new_user.email,
            'full_name': new_user.full_name,
            'role': new_user.role,
            'is_active': new_user.is_active
        }
    except Exception as e:
        db.rollback()
        raise e
    finally:
        db.close()


def update_user(user_id, updates):
    db = get_db()
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            return False

        for key, value in updates.items():
            if key != 'id':
                setattr(user, key, value)
        db.commit()
        return True
    except Exception as e:
        db.rollback()
        return False
    finally:
        db.close()


def delete_user(user_id):
    db = get_db()
    try:
        admin_count = db.query(User).filter(User.role == 'admin').count()
        user = db.query(User).filter(User.id == user_id).first()
        if user and user.role == 'admin' and admin_count <= 1:
            return False

        db.query(User).filter(User.id == user_id).delete()
        db.commit()
        return True
    except Exception as e:
        db.rollback()
        return False
    finally:
        db.close()


def authenticate_user(username, password):
    db = get_db()
    try:
        user = db.query(User).filter(User.username == username).first()
        if user and user.is_active and check_password_hash(user.password_hash, password):
            return {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'full_name': user.full_name,
                'role': user.role,
                'is_active': user.is_active
            }
        return None
    finally:
        db.close()


def update_last_login(user_id):
    db = get_db()
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if user:
            user.last_login = datetime.utcnow()
            db.commit()
    except Exception as e:
        db.rollback()
    finally:
        db.close()


def load_all_projects():
    db = get_db()
    try:
        projects = db.query(Project).all()
        return [{
            'id': p.id,
            'project_name': p.project_name,
            'user_id': p.user_id,
            'created_date': p.created_date.isoformat() if p.created_date else None,
            'modified_date': p.modified_date.isoformat() if p.modified_date else None,
            'implementation_items': p.implementation_items or [],
            'logistics_items': p.logistics_items or [],
            'software_items': p.software_items or [],
            'annual_maintenance_cost': p.annual_maintenance_cost,
            'expected_annual_benefit': p.expected_annual_benefit,
            'project_lifetime_years': p.project_lifetime_years,
            'discount_rate': p.discount_rate,
            'markup_percentages': p.markup_percentages or {},
            'currency_code': p.currency_code,
            'currency_symbol': p.currency_symbol,
            'exchange_rate': p.exchange_rate,
            'approval_status': p.approval_status,
            'approved_by': p.approved_by,
            'approved_date': p.approved_date.isoformat() if p.approved_date else None,
            'approval_comments': p.approval_comments
        } for p in projects]
    finally:
        db.close()


def save_all_projects(projects_list):
    db = get_db()
    try:
        db.query(Project).delete()

        for project_data in projects_list:
            created_date = None
            if project_data.get('created_date'):
                try:
                    created_date = datetime.fromisoformat(project_data['created_date'])
                except:
                    created_date = datetime.utcnow()

            modified_date = None
            if project_data.get('modified_date'):
                try:
                    modified_date = datetime.fromisoformat(project_data['modified_date'])
                except:
                    modified_date = datetime.utcnow()

            approved_date = None
            if project_data.get('approved_date'):
                try:
                    approved_date = datetime.fromisoformat(project_data['approved_date'])
                except:
                    pass

            project = Project(
                id=project_data.get('id', str(uuid.uuid4())),
                project_name=project_data.get('project_name', 'Unnamed Project'),
                user_id=project_data.get('user_id'),
                created_date=created_date or datetime.utcnow(),
                modified_date=modified_date or datetime.utcnow(),
                implementation_items=project_data.get('implementation_items', []),
                logistics_items=project_data.get('logistics_items', []),
                software_items=project_data.get('software_items', []),
                annual_maintenance_cost=project_data.get('annual_maintenance_cost', 0.0),
                expected_annual_benefit=project_data.get('expected_annual_benefit', 0.0),
                project_lifetime_years=project_data.get('project_lifetime_years', 3.0),
                discount_rate=project_data.get('discount_rate', 10.0),
                markup_percentages=project_data.get('markup_percentages', {}),
                currency_code=project_data.get('currency_code', 'USD'),
                currency_symbol=project_data.get('currency_symbol', '$'),
                exchange_rate=project_data.get('exchange_rate', 1.0),
                approval_status=project_data.get('approval_status', 'pending'),
                approved_by=project_data.get('approved_by'),
                approved_date=approved_date,
                approval_comments=project_data.get('approval_comments', '')
            )
            db.add(project)
        db.commit()
    except Exception as e:
        db.rollback()
        raise e
    finally:
        db.close()


def get_project_by_id(project_id):
    db = get_db()
    try:
        project = db.query(Project).filter(Project.id == project_id).first()
        if project:
            return {
                'id': project.id,
                'project_name': project.project_name,
                'user_id': project.user_id,
                'created_date': project.created_date.isoformat() if project.created_date else None,
                'modified_date': project.modified_date.isoformat() if project.modified_date else None,
                'implementation_items': project.implementation_items or [],
                'logistics_items': project.logistics_items or [],
                'software_items': project.software_items or [],
                'annual_maintenance_cost': project.annual_maintenance_cost,
                'expected_annual_benefit': project.expected_annual_benefit,
                'project_lifetime_years': project.project_lifetime_years,
                'discount_rate': project.discount_rate,
                'markup_percentages': project.markup_percentages or {},
                'currency_code': project.currency_code,
                'currency_symbol': project.currency_symbol,
                'exchange_rate': project.exchange_rate,
                'approval_status': project.approval_status,
                'approved_by': project.approved_by,
                'approved_date': project.approved_date.isoformat() if project.approved_date else None,
                'approval_comments': project.approval_comments
            }
        return None
    finally:
        db.close()


def save_project_to_db(project_data, user_id):
    db = get_db()
    try:
        now = datetime.utcnow()

        if 'id' not in project_data or not project_data['id']:
            project_data['id'] = str(uuid.uuid4())
            project_data['created_date'] = now.isoformat()
            project_data['approval_status'] = 'pending'
            project_data['approved_by'] = None
            project_data['approved_date'] = None
            project_data['approval_comments'] = ''

        project_data['modified_date'] = now.isoformat()
        project_data['user_id'] = user_id

        project_data.setdefault('expected_annual_benefit', 0)
        project_data.setdefault('project_lifetime_years', 3)
        project_data.setdefault('discount_rate', 10)
        project_data.setdefault('currency_code', 'USD')
        project_data.setdefault('currency_symbol', '$')
        project_data.setdefault('exchange_rate', 1.0)

        for item in project_data.get('implementation_items', []):
            item.setdefault('rate', 0)
            item.setdefault('no_of_resource', 1)
            item.setdefault('allocation', 100)
            item.setdefault('days', 1)
            item.setdefault('months', 1)
            item.setdefault('role', '')
            item.setdefault('location', '')
            item.setdefault('workplace_type', 'Remote')

        for item in project_data.get('software_items', []):
            item.setdefault('qty', 1)
            item.setdefault('unit_price', 0)
            item.setdefault('years', 1)
            item.setdefault('description', '')

        for item in project_data.get('logistics_items', []):
            item.setdefault('no_of_resources', 1)
            item.setdefault('round_trips', 0)
            item.setdefault('days_onsite', 0)
            item.setdefault('per_diem', 0)
            item.setdefault('accommodation', 0)
            item.setdefault('transport_per_day', 0)
            item.setdefault('flight_cost', 0)
            item.setdefault('role', '')
            item.setdefault('consultant_name', '')
            item.setdefault('location_base', '')

        db.query(Project).filter(Project.id == project_data['id']).delete()

        created_date = None
        if project_data.get('created_date'):
            try:
                created_date = datetime.fromisoformat(project_data['created_date'])
            except:
                created_date = now

        modified_date = now
        approved_date = None
        if project_data.get('approved_date'):
            try:
                approved_date = datetime.fromisoformat(project_data['approved_date'])
            except:
                pass

        project = Project(
            id=project_data['id'],
            project_name=project_data.get('project_name', 'Unnamed Project'),
            user_id=user_id,
            created_date=created_date or now,
            modified_date=modified_date,
            implementation_items=project_data.get('implementation_items', []),
            logistics_items=project_data.get('logistics_items', []),
            software_items=project_data.get('software_items', []),
            annual_maintenance_cost=project_data.get('annual_maintenance_cost', 0.0),
            expected_annual_benefit=project_data.get('expected_annual_benefit', 0.0),
            project_lifetime_years=project_data.get('project_lifetime_years', 3.0),
            discount_rate=project_data.get('discount_rate', 10.0),
            markup_percentages=project_data.get('markup_percentages', {}),
            currency_code=project_data.get('currency_code', 'USD'),
            currency_symbol=project_data.get('currency_symbol', '$'),
            exchange_rate=project_data.get('exchange_rate', 1.0),
            approval_status=project_data.get('approval_status', 'pending'),
            approved_by=project_data.get('approved_by'),
            approved_date=approved_date,
            approval_comments=project_data.get('approval_comments', '')
        )
        db.add(project)
        db.commit()

        return project_data
    except Exception as e:
        db.rollback()
        raise e
    finally:
        db.close()


def delete_project_from_db(project_id, user_id):
    db = get_db()
    try:
        project = db.query(Project).filter(Project.id == project_id).first()
        if not project:
            return False

        user = db.query(User).filter(User.id == user_id).first()
        if project.user_id == user_id or (user and user.role == 'admin'):
            db.query(Project).filter(Project.id == project_id).delete()
            db.commit()
            return True
        return False
    except Exception as e:
        db.rollback()
        return False
    finally:
        db.close()


def get_users_for_project(project_id):
    db = get_db()
    try:
        assignments = db.query(UserProjectAssignment).filter(
            UserProjectAssignment.project_id == project_id
        ).all()
        return [a.user_id for a in assignments]
    finally:
        db.close()


def load_user_projects_assignment():
    db = get_db()
    try:
        assignments = db.query(UserProjectAssignment).all()
        result = {}
        for assignment in assignments:
            if assignment.project_id not in result:
                result[assignment.project_id] = []
            result[assignment.project_id].append(assignment.user_id)
        return result
    finally:
        db.close()


def save_user_projects_assignment(assignments):
    db = get_db()
    try:
        db.query(UserProjectAssignment).delete()
        for project_id, user_ids in assignments.items():
            for user_id in user_ids:
                assignment = UserProjectAssignment(
                    project_id=project_id,
                    user_id=user_id
                )
                db.add(assignment)
        db.commit()
    except Exception as e:
        db.rollback()
        raise e
    finally:
        db.close()


# ===================== HELPER FUNCTIONS =====================

def get_currency_symbol(currency_code):
    return SUPPORTED_CURRENCIES.get(currency_code, {}).get('symbol', '$')


def get_decimal_places(currency_code):
    return SUPPORTED_CURRENCIES.get(currency_code, {}).get('decimal_places', 2)


def format_currency(value, currency_code='USD', currency_symbol=None):
    if currency_symbol is None:
        currency_symbol = get_currency_symbol(currency_code)
    decimal_places = get_decimal_places(currency_code)
    if decimal_places == 0:
        formatted_value = f"{safe_float(value):,.0f}"
    else:
        formatted_value = f"{safe_float(value):,.{decimal_places}f}"
    return f"{currency_symbol}{formatted_value}"


def safe_float(value):
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def safe_int(value):
    if value is None:
        return 0
    try:
        return int(value)
    except (ValueError, TypeError):
        return 0


def safe_str(value):
    if value is None:
        return ""
    return str(value)


def get_roi_category(roi_percentage: float) -> str:
    if roi_percentage >= 50:
        return 'Excellent'
    elif roi_percentage >= 25:
        return 'Good'
    elif roi_percentage >= 10:
        return 'Moderate'
    elif roi_percentage >= 0:
        return 'Low'
    return 'Negative'


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def convert_currency(amount, from_currency, to_currency):
    if from_currency == to_currency:
        return amount

    rates = load_currency_rates()

    if from_currency != 'USD':
        usd_rate = rates.get(from_currency, 1.0)
        if usd_rate > 0:
            amount_in_usd = amount / usd_rate
        else:
            amount_in_usd = amount
    else:
        amount_in_usd = amount

    if to_currency != 'USD':
        target_rate = rates.get(to_currency, 1.0)
        return amount_in_usd * target_rate
    else:
        return amount_in_usd


def convert_project_items(project_data, new_currency):
    if project_data.get('currency_code') == new_currency:
        return project_data

    old_currency = project_data.get('currency_code', 'USD')
    rates = load_currency_rates()

    if old_currency == 'USD':
        rate = rates.get(new_currency, 1.0)
    elif new_currency == 'USD':
        rate = 1.0 / rates.get(old_currency, 1.0)
    else:
        rate_to_usd = 1.0 / rates.get(old_currency, 1.0) if old_currency != 'USD' else 1.0
        rate_from_usd = rates.get(new_currency, 1.0)
        rate = rate_to_usd * rate_from_usd

    for item in project_data.get('implementation_items', []):
        if 'rate' in item:
            item['rate'] = item['rate'] * rate

    for item in project_data.get('software_items', []):
        if 'unit_price' in item:
            item['unit_price'] = item['unit_price'] * rate

    for item in project_data.get('logistics_items', []):
        if 'per_diem' in item:
            item['per_diem'] = item['per_diem'] * rate
        if 'accommodation' in item:
            item['accommodation'] = item['accommodation'] * rate
        if 'transport_per_day' in item:
            item['transport_per_day'] = item['transport_per_day'] * rate
        if 'flight_cost' in item:
            item['flight_cost'] = item['flight_cost'] * rate

    if 'annual_maintenance_cost' in project_data:
        project_data['annual_maintenance_cost'] = project_data['annual_maintenance_cost'] * rate
    if 'expected_annual_benefit' in project_data:
        project_data['expected_annual_benefit'] = project_data['expected_annual_benefit'] * rate

    project_data['currency_code'] = new_currency
    project_data['currency_symbol'] = get_currency_symbol(new_currency)
    project_data['exchange_rate'] = rate

    return project_data


# ===================== DECORATORS =====================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if BASE_PATH:
                return redirect(f'{BASE_PATH}/login')
            return redirect(url_for('login'))
        return f(*args, **kwargs)

    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if BASE_PATH:
                return redirect(f'{BASE_PATH}/login')
            return redirect(url_for('login'))
        user = get_user_by_id(session['user_id'])
        if not user or user.get('role') != 'admin':
            flash('Admin access required', 'danger')
            if BASE_PATH:
                return redirect(f'{BASE_PATH}/dashboard')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)

    return decorated_function


def can_view_all_projects(user_id: str) -> bool:
    user = get_user_by_id(user_id)
    if not user:
        return False
    return user.get('role') in ['admin', 'ceo', 'procurement', 'technical_lead', 'viewer']


def can_edit_project(user_id: str, project_id: str) -> bool:
    user = get_user_by_id(user_id)
    if not user:
        return False
    if user.get('role') in ['admin', 'procurement']:
        return True
    if user.get('role') in ['ceo', 'technical_lead', 'viewer']:
        return False
    project = get_project_by_id(project_id)
    if project and project.get('user_id') == user_id:
        return True
    return False


def can_delete_project(user_id: str, project_id: str) -> bool:
    user = get_user_by_id(user_id)
    if not user:
        return False
    return user.get('role') == 'admin'


def can_create_project(user_id: str) -> bool:
    user = get_user_by_id(user_id)
    if not user:
        return False
    return user.get('role') not in ['technical_lead', 'ceo', 'viewer']


def can_approve_project(user_id: str, project_id: str) -> bool:
    user = get_user_by_id(user_id)
    if not user:
        return False
    return user.get('role') in ['admin', 'ceo']


def can_download_templates(user_id: str) -> bool:
    user = get_user_by_id(user_id)
    if not user:
        return False
    if user.get('role') == 'technical_lead':
        return True
    return user.get('role') in ['admin', 'procurement', 'user']


def get_report_type(user_id: str, project_id: str) -> str:
    user = get_user_by_id(user_id)
    if not user:
        return 'full'
    if user.get('role') == 'technical_lead':
        return 'executive_only'
    return 'full'


def can_user_access_project(user_id: str, project_id: str) -> bool:
    project = get_project_by_id(project_id)
    if not project:
        return False
    user = get_user_by_id(user_id)
    if user and user.get('role') in ['admin', 'ceo', 'procurement', 'technical_lead']:
        return True
    if project.get('user_id') == user_id:
        return True
    shared_users = get_users_for_project(project_id)
    return user_id in shared_users


def get_accessible_projects(user_id: str) -> List[Dict]:
    all_projects = load_all_projects()
    user = get_user_by_id(user_id)
    if user and user.get('role') in ['admin', 'ceo', 'procurement', 'technical_lead']:
        return all_projects
    accessible = []
    for project in all_projects:
        if can_user_access_project(user_id, project.get('id')):
            accessible.append(project)
    return accessible


# ===================== DATA CLASSES =====================

@dataclass
class ImplementationCostItem:
    role: str = ""
    rate: float = 0.0
    no_of_resource: int = 1
    location: str = ""
    workplace_type: str = "Remote"
    allocation: float = 100.0
    days: float = 20.0
    months: float = 1.0

    def __post_init__(self):
        self.rate = safe_float(self.rate)
        self.no_of_resource = safe_int(self.no_of_resource)
        self.allocation = safe_float(self.allocation)
        self.days = safe_float(self.days)
        self.months = safe_float(self.months)
        self.role = safe_str(self.role)
        self.location = safe_str(self.location)
        self.workplace_type = safe_str(self.workplace_type)

    def total(self) -> float:
        return self.rate * self.no_of_resource * (self.allocation / 100) * self.days * self.months


@dataclass
class LogisticsCostItem:
    role: str = ""
    consultant_name: str = ""
    location_base: str = ""
    no_of_resources: int = 1
    round_trips: int = 0
    days_onsite: int = 0
    per_diem: float = 0.0
    accommodation: float = 0.0
    transport_per_day: float = 0.0
    flight_cost: float = 0.0

    def __post_init__(self):
        self.no_of_resources = safe_int(self.no_of_resources)
        self.round_trips = safe_int(self.round_trips)
        self.days_onsite = safe_int(self.days_onsite)
        self.per_diem = safe_float(self.per_diem)
        self.accommodation = safe_float(self.accommodation)
        self.transport_per_day = safe_float(self.transport_per_day)
        self.flight_cost = safe_float(self.flight_cost)
        self.role = safe_str(self.role)
        self.consultant_name = safe_str(self.consultant_name)
        self.location_base = safe_str(self.location_base)

    def total(self) -> float:
        flight_total = self.no_of_resources * self.round_trips * self.flight_cost
        transport_total = self.no_of_resources * self.days_onsite * self.transport_per_day
        per_diem_total = self.no_of_resources * self.days_onsite * self.per_diem
        accommodation_total = self.no_of_resources * self.days_onsite * self.accommodation
        return flight_total + transport_total + per_diem_total + accommodation_total


@dataclass
class SoftwareCostItem:
    description: str = ""
    qty: int = 1
    unit_price: float = 0.0
    years: float = 1.0

    def __post_init__(self):
        self.qty = safe_int(self.qty)
        self.unit_price = safe_float(self.unit_price)
        self.years = safe_float(self.years)
        self.description = safe_str(self.description)

    def total(self) -> float:
        return self.qty * self.unit_price * self.years


@dataclass
class ProjectCosting:
    id: str = ""
    project_name: str = "New Project"
    user_id: str = ""
    created_date: str = ""
    modified_date: str = ""
    implementation_items: List[Dict] = field(default_factory=list)
    logistics_items: List[Dict] = field(default_factory=list)
    software_items: List[Dict] = field(default_factory=list)
    annual_maintenance_cost: float = 0.0
    expected_annual_benefit: float = 0.0
    project_lifetime_years: float = 3.0
    discount_rate: float = 10.0
    markup_percentages: Dict[str, float] = field(default_factory=lambda: {
        "implementation": 0, "software": 0, "logistics": 0, "maintenance": 0
    })
    currency_code: str = "USD"
    currency_symbol: str = "$"
    exchange_rate: float = 1.0
    approval_status: str = "pending"
    approved_by: str = None
    approved_date: str = None
    approval_comments: str = ""

    def __post_init__(self):
        self.annual_maintenance_cost = safe_float(self.annual_maintenance_cost)
        self.expected_annual_benefit = safe_float(self.expected_annual_benefit)
        self.project_lifetime_years = safe_float(self.project_lifetime_years)
        self.discount_rate = safe_float(self.discount_rate)
        self.exchange_rate = safe_float(self.exchange_rate)
        if self.markup_percentages:
            for key in self.markup_percentages:
                self.markup_percentages[key] = safe_float(self.markup_percentages[key])

    def calculate_totals(self) -> Dict[str, float]:
        impl_total = sum(ImplementationCostItem(**item).total() for item in self.implementation_items)
        log_total = sum(LogisticsCostItem(**item).total() for item in self.logistics_items)
        sw_total = sum(SoftwareCostItem(**item).total() for item in self.software_items)
        grand_total = impl_total + log_total + sw_total + self.annual_maintenance_cost
        return {
            "implementation": impl_total,
            "logistics": log_total,
            "software": sw_total,
            "maintenance": self.annual_maintenance_cost,
            "grand_total": grand_total
        }

    def calculate_totals_with_markup(self) -> Dict[str, float]:
        impl_base = sum(ImplementationCostItem(**item).total() for item in self.implementation_items)
        log_base = sum(LogisticsCostItem(**item).total() for item in self.logistics_items)
        sw_base = sum(SoftwareCostItem(**item).total() for item in self.software_items)
        maint_base = self.annual_maintenance_cost

        impl_markup = impl_base * (1 + self.markup_percentages.get("implementation", 0) / 100)
        log_markup = log_base * (1 + self.markup_percentages.get("logistics", 0) / 100)
        sw_markup = sw_base * (1 + self.markup_percentages.get("software", 0) / 100)
        maint_markup = maint_base * (1 + self.markup_percentages.get("maintenance", 0) / 100)

        return {
            "implementation_base": impl_base,
            "implementation_markup": impl_markup,
            "implementation_percent": self.markup_percentages.get("implementation", 0),
            "logistics_base": log_base,
            "logistics_markup": log_markup,
            "logistics_percent": self.markup_percentages.get("logistics", 0),
            "software_base": sw_base,
            "software_markup": sw_markup,
            "software_percent": self.markup_percentages.get("software", 0),
            "maintenance_base": maint_base,
            "maintenance_markup": maint_markup,
            "maintenance_percent": self.markup_percentages.get("maintenance", 0),
            "grand_total": impl_markup + log_markup + sw_markup + maint_markup
        }


def calculate_roi_safe(project_data: Dict) -> Dict:
    try:
        if not project_data:
            return {
                'total_base_cost': 0, 'total_cost_with_markup': 0, 'total_markup_amount': 0,
                'expected_annual_benefit': 0, 'project_lifetime_years': 3,
                'discount_rate': 10, 'simple_roi': 0, 'npv': 0, 'payback_period': 0,
                'benefit_cost_ratio': 0, 'roi_category': 'Low', 'weighted_markup': 0
            }

        costing = ProjectCosting(**project_data)
        totals_with_markup = costing.calculate_totals_with_markup()

        total_base_cost = (
                totals_with_markup.get('implementation_base', 0) +
                totals_with_markup.get('logistics_base', 0) +
                totals_with_markup.get('software_base', 0) +
                totals_with_markup.get('maintenance_base', 0)
        )

        total_cost_with_markup = totals_with_markup.get('grand_total', 0)
        total_markup_amount = total_cost_with_markup - total_base_cost

        markups = project_data.get('markup_percentages', {
            "implementation": 0, "software": 0, "logistics": 0, "maintenance": 0
        })

        weighted_markup = 0
        if total_base_cost > 0:
            weighted_markup = (
                                      (totals_with_markup.get('implementation_base', 0) * markups.get("implementation",
                                                                                                      0)) +
                                      (totals_with_markup.get('software_base', 0) * markups.get("software", 0)) +
                                      (totals_with_markup.get('logistics_base', 0) * markups.get("logistics", 0)) +
                                      (totals_with_markup.get('maintenance_base', 0) * markups.get("maintenance", 0))
                              ) / total_base_cost

        expected_annual_benefit = project_data.get('expected_annual_benefit', 0)
        project_lifetime_years = project_data.get('project_lifetime_years', 3)
        discount_rate = project_data.get('discount_rate', 10) / 100

        if expected_annual_benefit > 0:
            total_benefit = expected_annual_benefit * project_lifetime_years
            if total_base_cost > 0:
                simple_roi = ((total_benefit / total_base_cost) - 1) * 100
            else:
                simple_roi = 0

            npv = -total_base_cost
            for year in range(1, int(project_lifetime_years) + 1):
                npv += expected_annual_benefit / ((1 + discount_rate) ** year)

            payback_period = total_base_cost / expected_annual_benefit if expected_annual_benefit > 0 else 0
            benefit_cost_ratio = total_benefit / total_base_cost if total_base_cost > 0 else 0
        else:
            if total_markup_amount > 0 and project_lifetime_years > 0:
                annual_benefit_from_markup = total_markup_amount / project_lifetime_years
            else:
                annual_benefit_from_markup = 0

            if total_base_cost > 0 and annual_benefit_from_markup > 0:
                total_benefit = annual_benefit_from_markup * project_lifetime_years
                simple_roi = ((total_benefit / total_base_cost) - 1) * 100
            else:
                simple_roi = 0

            npv = -total_base_cost
            for year in range(1, int(project_lifetime_years) + 1):
                npv += annual_benefit_from_markup / ((1 + discount_rate) ** year)

            payback_period = total_base_cost / annual_benefit_from_markup if annual_benefit_from_markup > 0 else 0
            benefit_cost_ratio = (
                                         annual_benefit_from_markup * project_lifetime_years) / total_base_cost if total_base_cost > 0 else 0

        simple_roi = max(simple_roi, 0)
        npv = max(npv, 0)

        return {
            'total_base_cost': total_base_cost,
            'total_cost_with_markup': total_cost_with_markup,
            'total_markup_amount': total_markup_amount,
            'expected_annual_benefit': expected_annual_benefit,
            'project_lifetime_years': project_lifetime_years,
            'discount_rate': discount_rate * 100,
            'simple_roi': simple_roi,
            'npv': npv,
            'payback_period': payback_period,
            'benefit_cost_ratio': benefit_cost_ratio,
            'roi_category': get_roi_category(simple_roi),
            'weighted_markup': weighted_markup
        }
    except Exception as e:
        print(f"ROI calculation error: {e}")
        traceback.print_exc()
        return {
            'total_base_cost': 0, 'total_cost_with_markup': 0, 'total_markup_amount': 0,
            'expected_annual_benefit': 0, 'project_lifetime_years': 3,
            'discount_rate': 10, 'simple_roi': 0, 'npv': 0, 'payback_period': 0,
            'benefit_cost_ratio': 0, 'roi_category': 'Low', 'weighted_markup': 0
        }


def update_project_approval(project_id: str, user_id: str, status: str, comments: str = "") -> bool:
    projects = load_all_projects()
    user = get_user_by_id(user_id)
    for i, project in enumerate(projects):
        if project.get('id') == project_id:
            projects[i]['approval_status'] = status
            projects[i]['approved_by'] = user.get('username') if user else None
            projects[i]['approved_date'] = datetime.now().isoformat()
            projects[i]['approval_comments'] = comments
            save_all_projects(projects)
            return True
    return False


# ===================== TEMPLATE FUNCTIONS =====================

def generate_pdf_report(project_data: Dict, report_type: str = 'full') -> io.BytesIO:
    buffer = io.BytesIO()
    settings = load_settings()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
    styles = getSampleStyleSheet()

    company_style = ParagraphStyle('CompanyStyle', parent=styles['Normal'], fontSize=8,
                                   textColor=colors.HexColor('#666666'), alignment=TA_RIGHT)
    normal_style = ParagraphStyle('CustomNormal', parent=styles['Normal'], fontSize=10, spaceAfter=6)
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24,
                                 textColor=colors.HexColor('#1e3c72'), alignment=TA_CENTER, spaceAfter=30)
    heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=16,
                                   textColor=colors.HexColor('#2c3e50'), spaceAfter=12, spaceBefore=20)
    subheading_style = ParagraphStyle('CustomSubheading', parent=styles['Heading3'], fontSize=12,
                                      textColor=colors.HexColor('#7f8c8d'), spaceAfter=8)

    currency_code = project_data.get('currency_code', 'USD')
    currency_symbol = project_data.get('currency_symbol', '$')

    try:
        logo_path = os.path.join(UPLOAD_FOLDER, settings.get('logo_filename', ''))
        if os.path.exists(logo_path) and settings.get('logo_filename'):
            logo = Image(logo_path, width=1.2 * inch, height=0.8 * inch)
        else:
            logo = Paragraph("", normal_style)
    except:
        logo = Paragraph("", normal_style)

    company_details = Paragraph(
        f"<b>{settings['company_name']}</b><br/>{settings['company_address']}<br/>Tel: {settings['company_phone']} | Email: {settings['company_email']}<br/>Web: {settings['company_website']}",
        company_style
    )

    header_data = [[logo, company_details]]
    header_table = Table(header_data, colWidths=[1.5 * inch, 4 * inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
    ]))

    story = [header_table, Spacer(1, 0.2 * inch), Paragraph("<hr/>", normal_style), Spacer(1, 0.1 * inch)]
    story.append(Paragraph("Project Costing Report", title_style))
    story.append(Paragraph(f"{safe_str(project_data.get('project_name', 'Unnamed Project'))}", heading_style))
    story.append(Spacer(1, 0.2 * inch))
    story.append(Paragraph(f"Report Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}", subheading_style))
    story.append(Paragraph(f"Currency: {currency_code} ({currency_symbol})", subheading_style))

    if project_data.get('modified_date'):
        try:
            modified = datetime.fromisoformat(project_data['modified_date']).strftime("%B %d, %Y")
            story.append(Paragraph(f"Last Modified: {modified}", subheading_style))
        except:
            pass

    approval_status = project_data.get('approval_status', 'pending')
    status_color = '#27ae60' if approval_status == 'approved' else '#e74c3c' if approval_status == 'rejected' else '#f39c12'
    story.append(Paragraph(f"Approval Status: <font color='{status_color}'><b>{approval_status.upper()}</b></font>",
                           subheading_style))
    story.append(Spacer(1, 0.3 * inch))

    costing = ProjectCosting(**project_data)
    totals_with_markup = costing.calculate_totals_with_markup()

    story.append(Paragraph("Executive Summary", heading_style))
    summary_data = [
        ["Cost Category", f"Base Amount ({currency_code})", "Markup", f"Final Amount ({currency_code})"],
        ["Implementation", format_currency(totals_with_markup['implementation_base'], currency_code, currency_symbol),
         f"{totals_with_markup['implementation_percent']:.1f}%",
         format_currency(totals_with_markup['implementation_markup'], currency_code, currency_symbol)],
        ["Software", format_currency(totals_with_markup['software_base'], currency_code, currency_symbol),
         f"{totals_with_markup['software_percent']:.1f}%",
         format_currency(totals_with_markup['software_markup'], currency_code, currency_symbol)],
        ["Logistics", format_currency(totals_with_markup['logistics_base'], currency_code, currency_symbol),
         f"{totals_with_markup['logistics_percent']:.1f}%",
         format_currency(totals_with_markup['logistics_markup'], currency_code, currency_symbol)],
        ["Annual Maintenance", format_currency(totals_with_markup['maintenance_base'], currency_code, currency_symbol),
         f"{totals_with_markup['maintenance_percent']:.1f}%",
         format_currency(totals_with_markup['maintenance_markup'], currency_code, currency_symbol)],
        ["", "", "", ""],
        ["", "", "GRAND TOTAL", format_currency(totals_with_markup['grand_total'], currency_code, currency_symbol)]
    ]

    summary_table = Table(summary_data, colWidths=[2.2 * inch, 1.2 * inch, 0.8 * inch, 1.5 * inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3c72')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (3, -2), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f39c12')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('SPAN', (0, -1), (2, -1)),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ('LINEBELOW', (0, -1), (-1, -1), 1, colors.black),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.3 * inch))

    roi_data = calculate_roi_safe(project_data)
    story.append(Paragraph("Return on Investment (ROI) Analysis", heading_style))

    story.append(
        Paragraph(
            f"<b>Total Base Cost (Investment):</b> {format_currency(roi_data['total_base_cost'], currency_code, currency_symbol)}",
            normal_style))
    story.append(Paragraph(
        f"<b>Total Markup Amount (Profit):</b> {format_currency(roi_data['total_markup_amount'], currency_code, currency_symbol)}",
        normal_style))
    story.append(Paragraph(
        f"<b>Total Cost with Markup:</b> {format_currency(roi_data['total_cost_with_markup'], currency_code, currency_symbol)}",
        normal_style))
    story.append(Spacer(1, 0.1 * inch))

    markups = project_data.get('markup_percentages', {})
    if markups.get('implementation', 0) > 0 or markups.get('software', 0) > 0 or markups.get('logistics',
                                                                                             0) > 0 or markups.get(
            'maintenance', 0) > 0:
        story.append(Paragraph(
            f"<i>Markup percentages used: Implementation: {markups.get('implementation', 0)}%, Software: {markups.get('software', 0)}%, Logistics: {markups.get('logistics', 0)}%, Maintenance: {markups.get('maintenance', 0)}%</i>",
            normal_style))
        story.append(Spacer(1, 0.1 * inch))

    if roi_data.get('expected_annual_benefit', 0) > 0:
        story.append(Paragraph(
            f"<i>Using custom Expected Annual Benefit: {format_currency(roi_data['expected_annual_benefit'], currency_code, currency_symbol)}</i>",
            normal_style))
    else:
        annual_benefit_from_markup = roi_data['total_markup_amount'] / roi_data['project_lifetime_years'] if roi_data[
                                                                                                                 'project_lifetime_years'] > 0 else 0
        story.append(Paragraph(
            f"<i>Annual Benefit derived from markup: {format_currency(annual_benefit_from_markup, currency_code, currency_symbol)} (Total markup {format_currency(roi_data['total_markup_amount'], currency_code, currency_symbol)} spread over {roi_data['project_lifetime_years']:.0f} years)</i>",
            normal_style))
    story.append(Spacer(1, 0.1 * inch))

    roi_table_data = [
        ["Metric", "Value"],
        ["Total Base Cost (Investment)", format_currency(roi_data['total_base_cost'], currency_code, currency_symbol)],
        ["Total Markup Amount", format_currency(roi_data['total_markup_amount'], currency_code, currency_symbol)],
        ["Project Lifetime", f"{roi_data['project_lifetime_years']:.0f} years"],
        ["Discount Rate", f"{roi_data['discount_rate']:.1f}%"],
        ["Simple ROI", f"{roi_data['simple_roi']:.1f}%"],
        ["Net Present Value (NPV)", format_currency(roi_data['npv'], currency_code, currency_symbol)],
        ["Payback Period", f"{roi_data['payback_period']:.1f} years"],
        ["Benefit-Cost Ratio", f"{roi_data['benefit_cost_ratio']:.2f}"],
        ["ROI Category", roi_data['roi_category']]
    ]

    if roi_data.get('weighted_markup', 0) > 0:
        roi_table_data.insert(1, ["Weighted Avg Markup", f"{roi_data['weighted_markup']:.1f}%"])

    roi_table = Table(roi_table_data, colWidths=[2.5 * inch, 3 * inch])
    roi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(roi_table)
    story.append(Spacer(1, 0.3 * inch))

    if report_type == 'executive_only':
        story.append(Paragraph(
            "<i>This is an executive summary report. Detailed cost breakdowns are available upon request.</i>",
            normal_style))
        story.append(Spacer(1, 0.3 * inch))
        doc.build(story)
        buffer.seek(0)
        return buffer

    # Implementation items
    impl_items = project_data.get('implementation_items', [])
    if impl_items:
        story.append(Paragraph("Implementation Cost Details", heading_style))
        impl_table_data = [
            ["Role", f"Rate ({currency_code})", "Resources", "Location", "Workplace", "Alloc%", "Days", "Months",
             f"Total ({currency_code})"]]
        for item in impl_items:
            impl_total = ImplementationCostItem(**item).total()
            impl_table_data.append([
                safe_str(item.get('role', ''))[:20],
                format_currency(safe_float(item.get('rate', 0)), currency_code, currency_symbol),
                str(safe_int(item.get('no_of_resource', 1))),
                safe_str(item.get('location', ''))[:15],
                safe_str(item.get('workplace_type', ''))[:10],
                f"{safe_float(item.get('allocation', 100)):.0f}%",
                str(safe_float(item.get('days', 1))),
                str(safe_float(item.get('months', 1))),
                format_currency(impl_total, currency_code, currency_symbol)
            ])
        impl_subtotal = sum(ImplementationCostItem(**item).total() for item in impl_items)
        impl_table_data.append(
            ["", "", "", "", "", "", "", "SUBTOTAL:", format_currency(impl_subtotal, currency_code, currency_symbol)])
        impl_table = Table(impl_table_data,
                           colWidths=[1.2 * inch, 0.7 * inch, 0.6 * inch, 0.9 * inch, 0.8 * inch, 0.6 * inch,
                                      0.5 * inch, 0.6 * inch, 0.9 * inch])
        impl_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (1, 1), (-2, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f4f8')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        story.append(impl_table)
        story.append(Spacer(1, 0.2 * inch))

    # Software items
    sw_items = project_data.get('software_items', [])
    if sw_items:
        story.append(Paragraph("Software Cost Details", heading_style))
        sw_table_data = [
            ["Description", "Quantity", f"Unit Price ({currency_code})", "Years", f"Total ({currency_code})"]]
        for item in sw_items:
            sw_total = SoftwareCostItem(**item).total()
            sw_table_data.append([
                safe_str(item.get('description', ''))[:30],
                str(safe_int(item.get('qty', 1))),
                format_currency(safe_float(item.get('unit_price', 0)), currency_code, currency_symbol),
                str(safe_float(item.get('years', 1))),
                format_currency(sw_total, currency_code, currency_symbol)
            ])
        sw_subtotal = sum(SoftwareCostItem(**item).total() for item in sw_items)
        sw_table_data.append(["", "", "", "SUBTOTAL:", format_currency(sw_subtotal, currency_code, currency_symbol)])
        sw_table = Table(sw_table_data, colWidths=[2.5 * inch, 0.8 * inch, 1 * inch, 0.8 * inch, 1.2 * inch])
        sw_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (1, 1), (-2, -1), 'RIGHT'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f4f8')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        story.append(sw_table)
        story.append(Spacer(1, 0.2 * inch))

    # Logistics items
    log_items = project_data.get('logistics_items', [])
    if log_items:
        story.append(Paragraph("Logistics Cost Details", heading_style))
        log_table_data = [
            ["Role", "Consultant", "Base", "Resources", "Trips", "Days", "Per Diem", "Accomm", "Transport", "Flight",
             f"Total ({currency_code})"]]
        for item in log_items:
            log_total = LogisticsCostItem(**item).total()
            log_table_data.append([
                safe_str(item.get('role', ''))[:12],
                safe_str(item.get('consultant_name', ''))[:12],
                safe_str(item.get('location_base', ''))[:8],
                str(safe_int(item.get('no_of_resources', 1))),
                str(safe_int(item.get('round_trips', 0))),
                str(safe_int(item.get('days_onsite', 0))),
                format_currency(safe_float(item.get('per_diem', 0)), currency_code, currency_symbol),
                format_currency(safe_float(item.get('accommodation', 0)), currency_code, currency_symbol),
                format_currency(safe_float(item.get('transport_per_day', 0)), currency_code, currency_symbol),
                format_currency(safe_float(item.get('flight_cost', 0)), currency_code, currency_symbol),
                format_currency(log_total, currency_code, currency_symbol)
            ])
        log_subtotal = sum(LogisticsCostItem(**item).total() for item in log_items)
        log_table_data.append(["", "", "", "", "", "", "", "", "", "SUBTOTAL:",
                               format_currency(log_subtotal, currency_code, currency_symbol)])
        log_table = Table(log_table_data,
                          colWidths=[0.8 * inch, 0.8 * inch, 0.7 * inch, 0.6 * inch, 0.6 * inch, 0.6 * inch, 0.7 * inch,
                                     0.7 * inch, 0.8 * inch, 0.7 * inch, 0.8 * inch])
        log_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('ALIGN', (3, 1), (-2, -1), 'RIGHT'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f4f8')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
        ]))
        story.append(log_table)
        story.append(Spacer(1, 0.2 * inch))

    # Annual Maintenance
    story.append(Paragraph("Annual Maintenance", heading_style))
    maint_base = safe_float(project_data.get('annual_maintenance_cost', 0))
    maint_percent = totals_with_markup['maintenance_percent']
    maint_markup = totals_with_markup['maintenance_markup']
    story.append(Paragraph(f"Base Annual Maintenance: {format_currency(maint_base, currency_code, currency_symbol)}",
                           normal_style))
    story.append(Paragraph(f"Markup Applied: {maint_percent:.1f}%", normal_style))
    story.append(Paragraph(f"Final Annual Maintenance: {format_currency(maint_markup, currency_code, currency_symbol)}",
                           normal_style))
    story.append(Spacer(1, 0.3 * inch))
    story.append(Spacer(1, 0.5 * inch))
    story.append(Paragraph("_" * 60, normal_style))
    story.append(Paragraph(
        settings.get('report_footer', 'This report was automatically generated by the Project Costing System.'),
        normal_style))
    story.append(Paragraph(f"Report ID: {safe_str(project_data.get('id', 'N/A'))}", normal_style))

    doc.build(story)
    buffer.seek(0)
    return buffer


def create_implementation_template(currency_code='USD'):
    output = io.BytesIO()
    currency_symbol = get_currency_symbol(currency_code)
    df = pd.DataFrame(
        columns=[f'Role', f'Rate ({currency_symbol})', '# Resources', 'Location', 'Workplace Type', 'Allocation (%)',
                 'Days', 'Months'])
    df.loc[0] = ['', 0, 1, '', 'Remote', 100, 1, 1]
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Implementation', index=False)
        worksheet = writer.sheets['Implementation']
        header_fill = PatternFill(start_color="1e3c72", end_color="1e3c72", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    output.seek(0)
    return output


def create_software_template(currency_code='USD'):
    output = io.BytesIO()
    currency_symbol = get_currency_symbol(currency_code)
    df = pd.DataFrame(columns=['Description', 'Quantity', f'Unit Price ({currency_symbol})', 'Years'])
    df.loc[0] = ['', 1, 0, 1]
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Software', index=False)
        worksheet = writer.sheets['Software']
        header_fill = PatternFill(start_color="1e3c72", end_color="1e3c72", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    output.seek(0)
    return output


def create_logistics_template(currency_code='USD'):
    output = io.BytesIO()
    currency_symbol = get_currency_symbol(currency_code)
    df = pd.DataFrame(
        columns=['Role', 'Consultant Name', 'Location Base', '# Resources', '# Round Trips', '# Days Onsite',
                 f'Per Diem ({currency_symbol}/day)', f'Accommodation ({currency_symbol}/day)',
                 f'Avg Transport/day ({currency_symbol})', f'Flight Cost ({currency_symbol})'])
    df.loc[0] = ['', '', '', 1, 0, 0, 0, 0, 0, 0]
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Logistics', index=False)
        worksheet = writer.sheets['Logistics']
        header_fill = PatternFill(start_color="1e3c72", end_color="1e3c72", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    output.seek(0)
    return output


def export_to_excel(project_data: Dict, tab_type: str):
    output = io.BytesIO()
    currency_code = project_data.get('currency_code', 'USD')
    currency_symbol = project_data.get('currency_symbol', '$')

    if tab_type == 'implementation':
        data = project_data.get('implementation_items', [])
        df = pd.DataFrame(data)
        if not df.empty:
            df = df[['role', 'rate', 'no_of_resource', 'location', 'workplace_type', 'allocation', 'days', 'months']]
            df.columns = ['Role', f'Rate ({currency_symbol})', '# Resources', 'Location', 'Workplace Type',
                          'Allocation (%)', 'Days', 'Months']
            df[f'Total ({currency_symbol})'] = df.apply(
                lambda x: safe_float(x[f'Rate ({currency_symbol})']) * safe_int(x['# Resources']) * (
                        safe_float(x['Allocation (%)']) / 100) * safe_float(x['Days']) * safe_float(x['Months']),
                axis=1)
    elif tab_type == 'software':
        data = project_data.get('software_items', [])
        df = pd.DataFrame(data)
        if not df.empty:
            df = df[['description', 'qty', 'unit_price', 'years']]
            df.columns = ['Description', 'Quantity', f'Unit Price ({currency_symbol})', 'Years']
            df[f'Total ({currency_symbol})'] = df['Quantity'] * df[f'Unit Price ({currency_symbol})'] * df['Years']
    elif tab_type == 'logistics':
        data = project_data.get('logistics_items', [])
        df = pd.DataFrame(data)
        if not df.empty:
            df = df[['role', 'consultant_name', 'location_base', 'no_of_resources', 'round_trips', 'days_onsite',
                     'per_diem', 'accommodation', 'transport_per_day', 'flight_cost']]
            df.columns = ['Role', 'Consultant Name', 'Location Base', '# Resources', '# Round Trips', '# Days Onsite',
                          f'Per Diem ({currency_symbol}/day)', f'Accommodation ({currency_symbol}/day)',
                          f'Avg Transport/day ({currency_symbol})', f'Flight Cost ({currency_symbol})']
            df[f'Total ({currency_symbol})'] = (
                    df['# Resources'] * df['# Round Trips'] * df[f'Flight Cost ({currency_symbol})'] + df[
                '# Resources'] * df[
                        '# Days Onsite'] * df[f'Avg Transport/day ({currency_symbol})'] + df['# Resources'] * df[
                        '# Days Onsite'] * df[
                        f'Per Diem ({currency_symbol}/day)'] + df['# Resources'] * df['# Days Onsite'] * df[
                        f'Accommodation ({currency_symbol}/day)'])
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=tab_type.capitalize(), index=False)
        worksheet = writer.sheets[tab_type.capitalize()]
        header_fill = PatternFill(start_color="1e3c72", end_color="1e3c72", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    output.seek(0)
    return output


def import_from_excel(file, tab_type: str) -> List[Dict]:
    df = pd.read_excel(file, sheet_name=0)
    items = []

    rate_col = None
    unit_price_col = None
    per_diem_col = None
    accommodation_col = None
    transport_col = None
    flight_cost_col = None

    for col in df.columns:
        col_lower = str(col).lower()
        if 'rate' in col_lower or ('price' in col_lower and tab_type == 'implementation'):
            rate_col = col
        elif 'unit price' in col_lower and tab_type == 'software':
            unit_price_col = col
        elif 'per diem' in col_lower and tab_type == 'logistics':
            per_diem_col = col
        elif 'accommodation' in col_lower and tab_type == 'logistics':
            accommodation_col = col
        elif 'transport' in col_lower and tab_type == 'logistics':
            transport_col = col
        elif 'flight cost' in col_lower and tab_type == 'logistics':
            flight_cost_col = col

    for _, row in df.iterrows():
        if tab_type == 'implementation':
            items.append({
                'role': safe_str(row.get('Role', '')),
                'rate': safe_float(row.get(rate_col, 0) if rate_col else row.get('Rate ($)', 0)),
                'no_of_resource': safe_int(row.get('# Resources', 1)),
                'location': safe_str(row.get('Location', '')),
                'workplace_type': safe_str(row.get('Workplace Type', 'Remote')),
                'allocation': safe_float(row.get('Allocation (%)', 100)),
                'days': safe_float(row.get('Days', 1)),
                'months': safe_float(row.get('Months', 1))
            })
        elif tab_type == 'software':
            items.append({
                'description': safe_str(row.get('Description', '')),
                'qty': safe_int(row.get('Quantity', 1)),
                'unit_price': safe_float(
                    row.get(unit_price_col, 0) if unit_price_col else row.get('Unit Price ($)', 0)),
                'years': safe_float(row.get('Years', 1))
            })
        elif tab_type == 'logistics':
            items.append({
                'role': safe_str(row.get('Role', '')),
                'consultant_name': safe_str(row.get('Consultant Name', '')),
                'location_base': safe_str(row.get('Location Base', '')),
                'no_of_resources': safe_int(row.get('# Resources', 1)),
                'round_trips': safe_int(row.get('# Round Trips', 0)),
                'days_onsite': safe_int(row.get('# Days Onsite', 0)),
                'per_diem': safe_float(row.get(per_diem_col, 0) if per_diem_col else row.get('Per Diem ($/day)', 0)),
                'accommodation': safe_float(
                    row.get(accommodation_col, 0) if accommodation_col else row.get('Accommodation ($/day)', 0)),
                'transport_per_day': safe_float(
                    row.get(transport_col, 0) if transport_col else row.get('Avg Transport/day ($)', 0)),
                'flight_cost': safe_float(
                    row.get(flight_cost_col, 0) if flight_cost_col else row.get('Flight Cost ($)', 0))
            })
    return items


# ===================== ROUTES =====================

@app.route('/api/currencies')
@login_required
def get_currencies():
    rates = load_currency_rates()
    currencies = []
    for code, info in SUPPORTED_CURRENCIES.items():
        currencies.append({
            'code': code,
            'symbol': info['symbol'],
            'name': info['name'],
            'rate_to_usd': rates.get(code, 1.0),
            'decimal_places': info['decimal_places']
        })
    return jsonify(currencies)


@app.route('/api/currencies/rates', methods=['GET', 'POST'])
@admin_required
def manage_currency_rates():
    if request.method == 'POST':
        rates = request.get_json()
        save_currency_rates(rates)
        return jsonify({"success": True, "message": "Exchange rates updated"})
    else:
        rates = load_currency_rates()
        return jsonify(rates)


@app.route('/api/projects/<project_id>/convert', methods=['POST'])
@login_required
def convert_project_currency(project_id):
    if not can_edit_project(session['user_id'], project_id):
        return jsonify({"error": "Not authorized to edit this project"}), 403

    data = request.get_json()
    new_currency = data.get('currency_code', 'USD')

    project = get_project_by_id(project_id)
    if not project:
        return jsonify({"error": "Project not found"}), 404

    converted_project = convert_project_items(project, new_currency)
    saved_project = save_project_to_db(converted_project, session['user_id'])

    return jsonify({
        "success": True,
        "message": f"Project converted to {new_currency}",
        "project": saved_project
    })


@app.route('/api/currencies/reset', methods=['POST'])
@admin_required
def reset_currency_rates():
    default_rates = {
        'USD': 1.0, 'EUR': 0.85, 'GBP': 0.73, 'JPY': 110.0,
        'CNY': 6.45, 'INR': 74.0, 'AUD': 1.35, 'CAD': 1.25,
        'CHF': 0.92, 'SGD': 1.35, 'HKD': 7.78, 'NZD': 1.43,
        'KRW': 1150.0, 'RUB': 73.5, 'BRL': 5.25, 'ZAR': 14.5,
        'AED': 3.67, 'SAR': 3.75, 'MYR': 4.15, 'THB': 33.0,
        'VND': 23000, 'IDR': 14400, 'PHP': 50.0, 'PKR': 160.0,
        'BDT': 85.0, 'NGN': 410.0, 'EGP': 15.7, 'TRY': 8.5,
        'MXN': 20.0, 'SEK': 8.5, 'NOK': 8.6, 'DKK': 6.3,
        'PLN': 3.8, 'CZK': 21.5, 'HUF': 300.0, 'ILS': 3.2,
        'CLP': 780.0, 'PEN': 3.8, 'COP': 3800.0
    }
    save_currency_rates(default_rates)
    return jsonify({"success": True, "message": "Exchange rates reset to defaults"})


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        user = authenticate_user(username, password)
        if user:
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['full_name'] = user.get('full_name', username)
            session['role'] = user.get('role', 'user')
            update_last_login(user['id'])
            flash(f'Welcome back, {session["full_name"]}!', 'success')
            if BASE_PATH:
                return redirect(f'{BASE_PATH}/dashboard')
            return redirect(url_for('dashboard'))
        flash('Invalid username or password', 'danger')
    return render_template('login.html', base_path=BASE_PATH)


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        email = request.form.get('email', '')
        full_name = request.form.get('full_name', '')
        if password != confirm_password:
            flash('Passwords do not match', 'danger')
            return render_template('register.html', base_path=BASE_PATH)
        if len(password) < 6:
            flash('Password must be at least 6 characters', 'danger')
            return render_template('register.html', base_path=BASE_PATH)
        user = create_user(username, password, email, full_name, 'user')
        if user:
            flash('Registration successful! Please login.', 'success')
            if BASE_PATH:
                return redirect(f'{BASE_PATH}/login')
            return redirect(url_for('login'))
        flash('Username already exists', 'danger')
    return render_template('register.html', base_path=BASE_PATH)


@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out', 'info')
    if BASE_PATH:
        return redirect(f'{BASE_PATH}/login')
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    user = get_user_by_id(session['user_id'])
    role = user.get('role', 'user') if user else 'user'
    settings = load_settings()
    return render_template('dashboard.html',
                           user=session.get('username'),
                           role=role,
                           settings=settings,
                           base_path=BASE_PATH)


@app.route('/admin')
@admin_required
def admin_panel():
    user = get_user_by_id(session['user_id'])
    role = user.get('role', 'user') if user else 'user'
    settings = load_settings()
    return render_template('admin.html',
                           user=session.get('username'),
                           role=role,
                           settings=settings,
                           base_path=BASE_PATH)


@app.route('/admin/settings', methods=['GET', 'POST'])
@admin_required
def admin_settings():
    settings = load_settings()
    if request.method == 'POST':
        settings['company_name'] = request.form.get('company_name', '')
        settings['company_address'] = request.form.get('company_address', '')
        settings['company_phone'] = request.form.get('company_phone', '')
        settings['company_email'] = request.form.get('company_email', '')
        settings['company_website'] = request.form.get('company_website', '')
        settings['system_name'] = request.form.get('system_name', '')
        settings['report_footer'] = request.form.get('report_footer', '')
        settings['default_currency'] = request.form.get('default_currency', 'USD')

        if 'logo' in request.files:
            file = request.files['logo']
            if file and file.filename != '' and allowed_file(file.filename):
                if settings.get('logo_filename'):
                    old_path = os.path.join(UPLOAD_FOLDER, settings['logo_filename'])
                    if os.path.exists(old_path):
                        os.remove(old_path)
                filename = f"logo_{datetime.now().strftime('%Y%m%d%H%M%S')}.{file.filename.rsplit('.', 1)[1].lower()}"
                file.save(os.path.join(UPLOAD_FOLDER, filename))
                settings['logo_filename'] = filename
        save_settings(settings)
        flash('Settings updated successfully!', 'success')
        if BASE_PATH:
            return redirect(f'{BASE_PATH}/admin/settings')
        return redirect(url_for('admin_settings'))

    currencies = [{'code': code, 'name': info['name'], 'symbol': info['symbol']}
                  for code, info in SUPPORTED_CURRENCIES.items()]
    return render_template('admin_settings.html',
                           user=session.get('username'),
                           role='admin',
                           settings=settings,
                           currencies=currencies,
                           base_path=BASE_PATH)


@app.route('/admin/currencies')
@admin_required
def admin_currencies():
    user = get_user_by_id(session['user_id'])
    role = user.get('role', 'user') if user else 'user'
    settings = load_settings()

    # Load rates with fallback
    try:
        rates = load_currency_rates()
        if rates is None:
            rates = {}
    except Exception as e:
        print(f"Error loading currency rates: {e}")
        rates = {}

    # Get default rates for comparison
    default_rates = get_default_exchange_rates()

    # If no rates in database, use defaults
    if not rates:
        rates = default_rates.copy()
        # Also add a last_updated timestamp
        rates['last_updated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    currencies = []
    for code, info in SUPPORTED_CURRENCIES.items():
        currencies.append({
            'code': code,
            'symbol': info['symbol'],
            'name': info['name'],
            'rate': rates.get(code, 1.0),
            'default_rate': default_rates.get(code, 1.0),
            'decimal_places': info['decimal_places']
        })
    currencies.sort(key=lambda x: x['code'])

    return render_template('admin_currencies.html',
                           user=session.get('username'),
                           role=role,
                           settings=settings,
                           currencies=currencies,
                           default_rates=default_rates,
                           rates=rates,  # ← CRITICAL
                           base_path=BASE_PATH)

@app.route('/admin/users')
@admin_required
def admin_get_users():
    users = load_users()
    for user in users:
        user.pop('password_hash', None)
    return jsonify(users)


@app.route('/admin/users', methods=['POST'])
@admin_required
def admin_create_user():
    data = request.get_json()
    username = data.get('username', '')
    password = data.get('password', '')
    email = data.get('email', '')
    full_name = data.get('full_name', '')
    role = data.get('role', 'user')
    if role not in ['admin', 'ceo', 'procurement', 'technical_lead', 'viewer', 'user']:
        return jsonify({"error": "Invalid role"}), 400
    if len(password) < 6:
        return jsonify({"error": "Password must be at least 6 characters"}), 400
    user = create_user(username, password, email, full_name, role)
    if user:
        user.pop('password_hash', None)
        return jsonify(user)
    return jsonify({"error": "Username already exists"}), 400


@app.route('/admin/users/<user_id>', methods=['PUT'])
@admin_required
def admin_update_user(user_id):
    data = request.get_json()
    updates = {}
    if 'email' in data:
        updates['email'] = data['email']
    if 'full_name' in data:
        updates['full_name'] = data['full_name']
    if 'role' in data:
        updates['role'] = data['role']
    if 'is_active' in data:
        updates['is_active'] = data['is_active']
    if 'password' in data and data['password']:
        if len(data['password']) >= 6:
            updates['password_hash'] = generate_password_hash(data['password'])
        else:
            return jsonify({"error": "Password must be at least 6 characters"}), 400
    if update_user(user_id, updates):
        return jsonify({"success": True})
    return jsonify({"error": "User not found"}), 404


@app.route('/admin/users/<user_id>', methods=['DELETE'])
@admin_required
def admin_delete_user(user_id):
    if user_id == session.get('user_id'):
        return jsonify({"error": "Cannot delete your own account"}), 400
    if delete_user(user_id):
        return jsonify({"success": True})
    return jsonify({"error": "Cannot delete the last admin account"}), 400


@app.route('/admin/projects')
@admin_required
def admin_get_projects():
    projects = load_all_projects()
    users = load_users()
    user_map = {u['id']: u for u in users}
    for project in projects:
        owner = user_map.get(project.get('user_id'))
        project['owner_name'] = owner.get('full_name') or owner.get('username') if owner else 'Unknown'
        project['shared_with'] = get_users_for_project(project.get('id'))
        project['approval_status'] = project.get('approval_status', 'pending')
    return jsonify(projects)


@app.route('/admin/projects/<project_id>/share', methods=['POST'])
@admin_required
def admin_share_project(project_id):
    data = request.get_json()
    user_ids = data.get('user_ids', [])
    assignments = load_user_projects_assignment()
    assignments[project_id] = user_ids
    save_user_projects_assignment(assignments)
    return jsonify({"success": True, "shared_with": user_ids})


@app.route('/admin/projects/<project_id>', methods=['DELETE'])
@admin_required
def admin_delete_project(project_id):
    projects = load_all_projects()
    projects = [p for p in projects if p.get('id') != project_id]
    save_all_projects(projects)
    assignments = load_user_projects_assignment()
    if project_id in assignments:
        del assignments[project_id]
        save_user_projects_assignment(assignments)
    return jsonify({"success": True})


@app.route('/api/projects/<project_id>/approve', methods=['POST'])
@login_required
def approve_project(project_id):
    if not can_approve_project(session['user_id'], project_id):
        return jsonify({"error": "Not authorized to approve projects"}), 403
    data = request.get_json()
    status = data.get('status', 'approved')
    comments = data.get('comments', '')
    if update_project_approval(project_id, session['user_id'], status, comments):
        return jsonify({"success": True, "status": status})
    return jsonify({"error": "Failed to update approval status"}), 400


@app.route('/')
@login_required
def index():
    user = get_user_by_id(session['user_id'])
    role = user.get('role', 'user') if user else 'user'
    if role in ['technical_lead', 'ceo', 'viewer']:
        flash(f'Access denied. {role.replace("_", " ").title()} role does not have permission to access Data Entry.',
              'danger')
        if BASE_PATH:
            return redirect(f'{BASE_PATH}/dashboard')
        return redirect(url_for('dashboard'))
    settings = load_settings()
    currencies = [{'code': code, 'symbol': info['symbol'], 'name': info['name']}
                  for code, info in SUPPORTED_CURRENCIES.items()]
    return render_template('index.html',
                           user=session.get('username'),
                           role=role,
                           settings=settings,
                           currencies=currencies,
                           base_path=BASE_PATH)


@app.route('/projects')
@login_required
def projects_page():
    user = get_user_by_id(session['user_id'])
    role = user.get('role', 'user') if user else 'user'
    settings = load_settings()
    currencies = [{'code': code, 'symbol': info['symbol'], 'name': info['name']}
                  for code, info in SUPPORTED_CURRENCIES.items()]
    return render_template('projects.html',
                           user=session.get('username'),
                           role=role,
                           settings=settings,
                           currencies=currencies,
                           base_path=BASE_PATH)


@app.route('/api/projects', methods=['GET'])
@login_required
def get_all_projects():
    try:
        projects = get_accessible_projects(session['user_id'])
        user = get_user_by_id(session['user_id'])

        result = []
        for project in projects:
            try:
                project_copy = {
                    'id': project.get('id', ''),
                    'project_name': project.get('project_name', 'Unnamed'),
                    'user_id': project.get('user_id', ''),
                    'created_date': project.get('created_date', ''),
                    'modified_date': project.get('modified_date', ''),
                    'implementation_items': project.get('implementation_items', []),
                    'software_items': project.get('software_items', []),
                    'logistics_items': project.get('logistics_items', []),
                    'annual_maintenance_cost': project.get('annual_maintenance_cost', 0),
                    'expected_annual_benefit': project.get('expected_annual_benefit', 0),
                    'project_lifetime_years': project.get('project_lifetime_years', 3),
                    'discount_rate': project.get('discount_rate', 10),
                    'markup_percentages': project.get('markup_percentages', {
                        "implementation": 0, "software": 0, "logistics": 0, "maintenance": 0
                    }),
                    'currency_code': project.get('currency_code', 'USD'),
                    'currency_symbol': project.get('currency_symbol', '$'),
                    'exchange_rate': project.get('exchange_rate', 1.0),
                    'approval_status': project.get('approval_status', 'pending'),
                    'can_edit': can_edit_project(session['user_id'], project.get('id')),
                    'can_delete': can_delete_project(session['user_id'], project.get('id')),
                    'can_approve': can_approve_project(session['user_id'], project.get('id')),
                    'user_role': user.get('role') if user else 'user'
                }

                roi_data = calculate_roi_safe(project)
                project_copy['roi'] = roi_data['simple_roi']
                project_copy['roi_category'] = roi_data['roi_category']
                project_copy['weighted_markup'] = roi_data['weighted_markup']

                result.append(project_copy)
            except Exception as project_err:
                print(f"Error processing project: {project_err}")
                continue

        result.sort(key=lambda x: x.get('modified_date', ''), reverse=True)
        return jsonify(result)
    except Exception as e:
        print(f"ERROR in get_all_projects: {str(e)}")
        traceback.print_exc()
        return jsonify([])


@app.route('/api/projects/<project_id>', methods=['GET'])
@login_required
def get_project(project_id):
    try:
        if can_user_access_project(session['user_id'], project_id):
            project = get_project_by_id(project_id)
            if project:
                project_copy = dict(project)
                project_copy['can_edit'] = can_edit_project(session['user_id'], project_id)
                project_copy['can_delete'] = can_delete_project(session['user_id'], project_id)
                project_copy['can_approve'] = can_approve_project(session['user_id'], project_id)

                roi_data = calculate_roi_safe(project)
                project_copy['roi'] = roi_data['simple_roi']
                project_copy['roi_category'] = roi_data['roi_category']
                project_copy['weighted_markup'] = roi_data['weighted_markup']

                return jsonify(project_copy)
        return jsonify({"error": "Project not found"}), 404
    except Exception as e:
        print(f"Error getting project: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/projects', methods=['POST'])
@login_required
def create_update_project():
    if not can_create_project(session['user_id']):
        return jsonify({"error": "Your role cannot create or edit projects"}), 403
    saved_project = save_project_to_db(request.get_json(), session['user_id'])
    return jsonify(saved_project)


@app.route('/api/projects/<project_id>', methods=['DELETE'])
@login_required
def delete_project(project_id):
    if can_delete_project(session['user_id'], project_id):
        if delete_project_from_db(project_id, session['user_id']):
            return jsonify({"status": "deleted", "id": project_id})
    return jsonify({"error": "Cannot delete project"}), 403


@app.route('/calculate', methods=['POST'])
@login_required
def calculate():
    data = request.get_json()
    costing = ProjectCosting(
        project_name=data.get('project_name', ''),
        implementation_items=data.get('implementation_items', []),
        logistics_items=data.get('logistics_items', []),
        software_items=data.get('software_items', []),
        annual_maintenance_cost=float(data.get('annual_maintenance_cost', 0))
    )
    return jsonify(costing.calculate_totals())


@app.route('/save_project', methods=['POST'])
@login_required
def save_project_route():
    if not can_create_project(session['user_id']):
        return jsonify({"error": "Your role cannot save projects"}), 403
    saved = save_project_to_db(request.get_json(), session['user_id'])
    return jsonify({"status": "saved", "project": saved.get('project_name'), "id": saved.get('id')})


@app.route('/report/<project_id>')
@login_required
def generate_report(project_id):
    if not can_user_access_project(session['user_id'], project_id):
        return jsonify({"error": "Access denied"}), 403
    project = get_project_by_id(project_id)
    if not project:
        return jsonify({"error": "Project not found"}), 404
    try:
        report_type = get_report_type(session['user_id'], project_id)
        pdf_buffer = generate_pdf_report(project, report_type)
        project_name = safe_str(project.get('project_name', 'project')).replace(' ', '_')
        return send_file(pdf_buffer, mimetype='application/pdf', as_attachment=True,
                         download_name=f"{project_name}_costing_report.pdf")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/export/template/<tab_type>')
@login_required
def export_template(tab_type):
    if not can_download_templates(session['user_id']):
        flash('Access denied', 'danger')
        if BASE_PATH:
            return redirect(f'{BASE_PATH}/dashboard')
        return redirect(url_for('dashboard'))

    currency_code = request.args.get('currency', 'USD')

    templates = {
        'implementation': lambda: create_implementation_template(currency_code),
        'software': lambda: create_software_template(currency_code),
        'logistics': lambda: create_logistics_template(currency_code)
    }
    if tab_type not in templates:
        return jsonify({"error": "Invalid tab type"}), 400
    file = templates[tab_type]()
    return send_file(file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'{tab_type}_template_{currency_code}.xlsx')


@app.route('/export/<tab_type>', methods=['POST'])
@login_required
def export_data(tab_type):
    file = export_to_excel(request.get_json(), tab_type)
    project_data = request.get_json()
    currency_code = project_data.get('currency_code', 'USD')
    return send_file(file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'{project_data.get("project_name", "project")}_{tab_type}_{currency_code}.xlsx')


@app.route('/import/<tab_type>', methods=['POST'])
@login_required
def import_data(tab_type):
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400
    try:
        items = import_from_excel(file, tab_type)
        return jsonify({"items": items, "count": len(items)})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


# ===================== MAIN =====================

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug_mode = not IS_RENDER
    app.run(host='0.0.0.0', port=port, debug=debug_mode)